from tqdm import tqdm
from openpyxl import load_workbook

from config_xlsx import COMPLEMENTO_CLIENTE_XLSX
from validaciones_texto import validarEncabezadoXlsx, setearFechaCelda

LOG_COMPLEMENTO_CLIENTE = dict()

def extraerComplementoCliente(registrosLog, pathXlsxEntrada):
    # archivo = '.\COMPLEMENTO_CLIENTE\COMPLEMENTO CLIENT vLite 202101.xlsx'
    archivo = pathXlsxEntrada
    registrosLog += 1
    LOG_COMPLEMENTO_CLIENTE.setdefault(registrosLog, {'INICIO_COMPLEMENTO_CLIENTE': 'Iniciando proceso de lectura del Archivo: %s' % archivo})
    try:
        encabezadoXls = COMPLEMENTO_CLIENTE_XLSX['ENCABEZADO']
        celda = COMPLEMENTO_CLIENTE_XLSX['COLUMNAS']
        coordenadaEcabezado = COMPLEMENTO_CLIENTE_XLSX['COORDENADA_ENCABEZADO']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        complementoCliente = dict()

        registrosLog += 1
        i = 0
        LOG_COMPLEMENTO_CLIENTE.setdefault(registrosLog , {'ENCABEZADO_COMPLEMENTOCLIENTE': 'Encabezado del Archivo: %s OK' % archivo})
        for fila in tqdm(iterable=hoja.rows, total= len(tuple(hoja.rows)), desc= 'Leyendo ComplementoCliente' , unit=' Fila'):

            i += 1
            if i >= 2:

                nroPoliza = int(fila[celda['NRO_POLIZA']].value)
                estadoPoliza = str(fila[celda['ESTADO_POLIZA']].value)
                fecUltPago = None
                fecMandato = None
                if fila[celda['FEC_ULT_PAG']].value is not None:
                    fecUltPago = setearFechaCelda(fila[celda['FEC_ULT_PAG']])
                if fila[celda['FECHA_MANDATO']].value is not None:
                    fecMandato = setearFechaCelda(fila[celda['FECHA_MANDATO']])
                complementoCliente[nroPoliza] = {'NRO_CERT': fila[celda['NRO_CERT']].value, 'ESTADO_POLIZA': estadoPoliza, 'FEC_ULT_PAG': fecUltPago, 'ESTADO_MANDATO': fila[celda['ESTADO_MANDATO']].value, 'FECHA_MANDATO': fecMandato}

        registrosLog += 1
        LOG_COMPLEMENTO_CLIENTE.setdefault(registrosLog, {'LECTURA_COMPLEMENTOCLIENTE': 'Lectura del Archivo: %s Finalizado - %s Filas' % (archivo, len(tuple(hoja.rows)))})
        return complementoCliente

    except Exception as e:
        registrosLog += 1
        errorMsg = 'Error al leer archivo;%s | %s' % (archivo, e)
        LOG_COMPLEMENTO_CLIENTE.setdefault(registrosLog, {'LECTURA_COMPLEMENTOCLIENTE': errorMsg})
        raise