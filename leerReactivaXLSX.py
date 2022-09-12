import datetime
import traceback

from openpyxl import load_workbook
from tqdm import tqdm

from complementoCliente import (LOG_COMPLEMENTO_CLIENTE,
                                extraerComplementoCliente)
from conexio_db import conectorDB
from config_xlsx import PATH_XLSX, REACTIVA_CONFIG_XLSX
from diccionariosDB import (buscarRutEjecutivosDb,
                            listaEstadoRetencionReactiva,
                            listaEstadoUtContacto, listaEstadoUtNoContacto,
                            periodoCampanasReactiva, CampanasPorPeriodoReactiva, estadoRetencionReacDesc)

from validaciones_texto import (convertirDataReact,
                                convertirListaReactiva,
                                formatearFechaMesAnterior, formatearIdCliente,
                                formatearNumeroPoliza, formatearRutGion,
                                primerDiaMes, setearCampanasReactiva,
                                setearCelda, setearCelda2, setearFechaCelda,
                                setearFechaInput, ultimoDiaMes,
                                validarEncabezadoXlsx, fechaUnida)

LOG_PROCESO_REACTIVA = dict()
campanasPorEjecutivos = dict()
clienteDuplicadoContactado = dict()
dataSalidaXlsx = dict()
inbound = REACTIVA_CONFIG_XLSX['INBOUND_VALOR']
outbound = REACTIVA_CONFIG_XLSX['OUTBOUND_VALOR']
listaEstadoRetencion = listaEstadoRetencionReactiva()

def validarClienteContactoDuplicado(contacto, nivelContacto, pkCliente, pk):
    clienteUnico = 'SI'
    contactoCliente = 'NO CONTACTADO'
    if clienteDuplicadoContactado.get(pkCliente):
        clienteUnico = 'NO'
        pkExistente = clienteDuplicadoContactado[pkCliente]['PK']
        nivelContactoExistente = clienteDuplicadoContactado[pkCliente]['NIVEL_CONTACTO']
        dataSalidaXlsx[pkExistente]['CLIENTE_UNICO'] = clienteUnico
        if contacto == 1:
            if nivelContacto <= nivelContactoExistente:
                contactoCliente = 'CONTACTADO'
                dataSalidaXlsx[pkExistente]['CONTACTO_CLIENTE'] = 'NO CONTACTADO'
                clienteDuplicadoContactado[pkCliente]['PK'] = pk
                clienteDuplicadoContactado[pkCliente]['NIVEL_CONTACTO'] = nivelContacto
    else:
        if contacto == 1:
            contactoCliente = 'CONTACTADO'
        clienteDuplicadoContactado[pkCliente] = {'PK' : pk, 'NIVEL_CONTACTO': nivelContacto}
    return contactoCliente, clienteUnico

def extraerBaseCertificacion(archivoCertificacionXls, fechaInicioPeriodo, fechaFinMes):
    archivo = archivoCertificacionXls
    archivoBaseCertificacion = REACTIVA_CONFIG_XLSX['ARCHIVO_BASE_CERTIFICACION']
    LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'INICIO_BASE_CERTIFICACION': 'Iniciando proceso de lectura del Archivo: %s' % archivo})
    try:
        encabezadoXls = archivoBaseCertificacion['ENCABEZADO']
        celda = archivoBaseCertificacion['COLUMNAS']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        baseCertificado = dict()
        i = 0
        validarArchivo = validarEncabezadoXlsx(hoja['A1:X1'], encabezadoXls, archivo)

        if type(validarArchivo) is not dict:
            LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1 , {'ENCABEZADO_BASE_CERTIFICACION': 'Encabezado del Archivo: %s OK' % archivo})
            for fila in tqdm(iterable=hoja.rows, total= len(tuple(hoja.rows)), desc= 'Leyendo BaseCertificacion' , unit=' Fila'):

                i += 1
                if i >= 2:

                    numeroPoliza = fila[celda['NRO_POLIZA']].value
                    idEmpleado = fila[celda['ID_EMPLEADO']].value
                    canal = str(fila[celda['CANAL']].value)
                    tipoCertificacion = str(fila[celda['TIPO_CERTIFICACION']].value)
                    fechaLlamado = setearFechaCelda(fila[celda['FECHA_LLAMADO']])

                    if numeroPoliza is None:
                        celdaCoordenada = setearCelda2(fila[celda['NRO_POLIZA']],0)
                        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'POLIZA_NULO': '%s;Numero de poliza NULL BaseCertificado;%s' % (celdaCoordenada, numeroPoliza)})
                        continue

                    if tipoCertificacion.upper() != 'GRABACIÓN CERTIFICADA':
                        continue

                    if type(fechaLlamado) is not datetime.date:
                        valorErroneo = fila[celda['FECHA_LLAMADO']].value
                        celdaCoordenada = setearCelda2(fila[0:celda['FECHA_LLAMADO']+1], len(fila[0:celda['FECHA_LLAMADO']])-1, i)
                        mensaje = '%s;FECHA_LLAMADO no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'FECHA_LLAMADO': mensaje})
                        continue

                    if fechaLlamado < fechaInicioPeriodo or fechaLlamado > fechaFinMes:
                        continue
                    
                    pk = '{0}_{1}'.format(str(numeroPoliza), str(idEmpleado))

                    if baseCertificado.get(pk):
                        if baseCertificado[pk]['FECHA_LLAMADO'] >= fechaLlamado:
                            continue
                    baseCertificado[pk] = {'NRO_POLIZA': int(numeroPoliza), 'FECHA_LLAMADO': fechaLlamado, 'ID_EMPLEADO': str(idEmpleado), 'CANAL': canal, 'TIPO_CERTIFICACION': tipoCertificacion}

            LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'LECTURA_BASE_CERTIFICACION': 'Lectura del Archivo: %s Finalizado - %s Filas' % (archivo, len(tuple(hoja.rows)))})
            return baseCertificado
        else:
            LOG_PROCESO_REACTIVA.setdefault('ENCABEZADO_BASE_CERTIFICACION', validarArchivo)
            raise
    except Exception as e:
        errorMsg = 'Error al leer archivo;%s | %s' % (archivo, e)
        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'LECTURA_BASE_CERTIFICACION': errorMsg})
        raise

def validarEstadoReact(estadoRetencion, estado):
    if estadoRetencion == 'Mantiene su producto' or estadoRetencion is None and estado == 'Terminado con Exito':
        estadoValidoReact = listaEstadoRetencion.get('Terminado con Exito')
    elif estadoRetencion is None:
        estadoValidoReact = listaEstadoRetencion.get(estado)
    elif estadoRetencion != 'Mantiene su producto':
        estadoValidoReact = listaEstadoRetencion.get(estadoRetencion)
    return estadoValidoReact

def validarContactoReact(saliente, estadoRetencion, estado, estadoUt, listaEstadoContactado):

    contactoReact = 0
    nivelContacto = 0
    if saliente == inbound:
        contactoReact = 1
        nivelContacto = 1
    elif saliente == outbound:
        if estadoRetencion == 'Mantiene su producto':
            contactoReact = 1
            nivelContacto = 1
        elif estadoRetencion is None:
            if estadoUt is not None:
                if listaEstadoContactado.get(estadoUt):
                    contactoReact = 1
                    nivelContacto = 2
            elif estado != 'Sin Gestion':
                contactoReact = 1
                nivelContacto = 2
        elif estadoRetencion != 'Mantiene su producto':
            if estadoUt is None:
                if estado != 'Sin Gestion':
                    contactoReact = 1
                    nivelContacto = 3
            elif listaEstadoContactado.get(estadoUt):
                contactoReact = 1
                nivelContacto = 3
    return contactoReact, nivelContacto

def exitoRepetidoPk(numeroPoliza, polizaExitoRepetido, gestionReactTxt):
    pkSalida = 0
    for pkData in polizaExitoRepetido[numeroPoliza]:
        if gestionReactTxt[pkData]['EXITO_REPETIDO_REACT'] == 1:
            pkSalida = pkData
    return pkSalida

def aprobarActualizarRegistro(estado, estadoValidoReact, contactoReact, estadoValidoReactData, contactoReactData):
    controlCambioPk = False
    indiceCambio = None
    if estadoValidoReact == 1 and estadoValidoReactData != 1:
        controlCambioPk = True
        indiceCambio = 'EstadoRetencion'
    elif contactoReact == 1 and contactoReactData == 0:
        controlCambioPk = True
        indiceCambio = 'Contactabilidad'
    elif estado != 'Sin Gestion' and estadoValidoReactData == 4:
        controlCambioPk = True
        indiceCambio = 'Estado'
    return controlCambioPk, indiceCambio

def insertarPeriodoCampanaEjecutivos(campanasEjecutivos: dict, fechaProceso):
    try:
        db = conectorDB()
        cursor = db.cursor()
        ejecutivosExistentes = periodoCampanasReactiva(fechaProceso)
        periodoEjecutivos = convertirListaReactiva(campanasEjecutivos, ejecutivosExistentes, fechaProceso)
        if len(periodoEjecutivos) > 0:
            sql = """INSERT INTO reactiva_campanas_periodo_ejecutivos (id_ejecutivo, periodo) VALUES (?, ?);"""
            cursor.executemany(sql, periodoEjecutivos)
            db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al insertar Periodo de Ejecutivos: %s' % (e))
    finally:
        cursor.close()
        db.close()

def insertarCampanaEjecutivos(campanasEjecutivos: dict, fechaProceso):
    try:
        db = conectorDB()
        camapanasPeriodoEjecutivos = periodoCampanasReactiva(fechaProceso)
        campanasPorPeriodo = []
        cursor = db.cursor()

        for valores in campanasEjecutivos.values():
            for polizas in valores.values():
                idEjecutivo = polizas['ID_EMPLEADO']
                if camapanasPeriodoEjecutivos.get(idEjecutivo):
                    campanasPorPeriodo += setearCampanasReactiva(polizas, camapanasPeriodoEjecutivos[idEjecutivo]['ID'])

        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'INSERTAR_CAMPANAS_EJECUTIVOS': '-----------------------------------------------------' })
        campanasExistentes = CampanasPorPeriodoReactiva(fechaProceso)
        if limpiarTablaCamapanasEjecutivos(fechaProceso):
            LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'LIMPIAR_CAMAPAÑAS_EJECUTIVOS': 'EliminarCampanaEjecutivos;Se eliminaron {0} Camapaña(s) existentes'.format(campanasExistentes)})

        if len(campanasPorPeriodo) > 0:
            sql = """INSERT INTO reactiva_campanas_ejecutivos (id_periodo_ejecutivo, numero_poliza, estado_retencion, estado_ut, in_out, certificacion, exito_repetido, estado_poliza, estado_final) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);"""
            cursor.executemany(sql, campanasPorPeriodo)
            db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al insertar Campañas de Ejecutivos del Periodo: %s' % (e))
    finally:
        cursor.close()
        db.close()

def limpiarTablaCamapanasEjecutivos(fechaProceso):
    try:
        db = conectorDB()
        cursor = db.cursor()

        sql = """DELETE FROM reactiva_campanas_ejecutivos WHERE id_periodo_ejecutivo IN (SELECT id FROM reactiva_campanas_periodo_ejecutivos WHERE periodo = ?);"""
        cursor.execute(sql, fechaProceso)
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al eliminar Campañas de Ejecutivos existentes: %s' % (e))
    finally:
        cursor.close()
        db.close()

def agregarCampanasPorEjecutivo(idEmpleado, pk, valoresCampanas: dict):
    
    if campanasPorEjecutivos.get(idEmpleado):
        campanasPorEjecutivos[idEmpleado].setdefault(pk, valoresCampanas)
    else:
        campanasPorEjecutivos[idEmpleado] = {pk: valoresCampanas}
    return 1

def formatearSaliente(valorEntrada):
    valorSalida = outbound
    if bool(valorEntrada):
        valorSalida = inbound
    return valorSalida

def leerArchivoReactiva(archivoEntrada, periodo, fechaInicioEntrada, fechaFinEntrada, archivoCertificacionXls, archivoComplmentoCliente):

    try:
        archivoSalida = REACTIVA_CONFIG_XLSX['SALIDA_TXT']
        columna = REACTIVA_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        xls = load_workbook(archivoEntrada, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]

        fechaInicioPeriodo = setearFechaInput(fechaInicioEntrada)
        fechaFinPeriodo = setearFechaInput(fechaFinEntrada)
        fechaIncioMes = primerDiaMes(periodo)
        fechaFinMes = ultimoDiaMes(periodo)

        complementoCliente = extraerComplementoCliente(len(LOG_PROCESO_REACTIVA), archivoComplmentoCliente)
        LOG_PROCESO_REACTIVA.update(LOG_COMPLEMENTO_CLIENTE)
        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'DIVISOR_PROCESO': '-----------------------------------------------------'})

        baseCertificacion = extraerBaseCertificacion(archivoCertificacionXls, fechaInicioPeriodo, fechaFinMes)
        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'DIVISOR_PROCESO': '-----------------------------------------------------'})

        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'INICIO_LECTURA_REACTIVA': 'Iniciando proceso de lectura del Archivo: %s' % archivoEntrada})

        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'ENCABEZADO_REACTIVA': 'Encabezado del Archivo: %s OK' % archivoEntrada})
        campanaDescripcion = {inbound : 'Inbound', outbound: 'Outbound'}
        gestionReactTxt = dict()
        polizaExitoRepetido = dict()
        polizaReactTxt = dict()
        certificacionReactTxt = dict()
        certificacion = {0: "n.a", 1: "Grabación Certificada", 2: "No certificada"}
        exitoRepetido = {0: "No", 1: "Si"}
        estadoFinal = {0: "No Retenido/No Gestionado", 1: "Retenido"}
        

        ejecutivosExistentesDb = buscarRutEjecutivosDb(fechaFinMes, fechaIncioMes)
        listaEstadoRetencionTexto = estadoRetencionReacDesc()
        listaEstadoContactado = listaEstadoUtContacto()
        i = 0
        cantidadCampanasValidas = 0
        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'INICIO_CELDAS_REACTIVA': 'Iniciando lectura de Celdas del Archivo: %s' % archivoEntrada})

        for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo Reactiva' , unit=' Fila'):

            i += 1
            if i >= 2:

                salienteEntrada = fila[columna['LLAMADA_SALIENTE']].value
                estado = str(fila[columna['ESTADO']].value)
                estadoUt = fila[columna['ESTADO_ULTIMA_TAREA']].value
                estadoRetencion = fila[columna['ESTADO_RETENCION']].value
                numeroPoliza, numeroPolizaCertificado = formatearNumeroPoliza(fila[columna['NRO_POLIZA']].value)
                idEmpleado = str(fila[columna['ID_EMPLEADO']].value)
                campanaId = str(fila[columna['CAMAPAÑA_ID']].value)
                campanaEntrada = str(fila[columna['NOMBRE_DE_CAMPAÑA']].value)
                polizasCampana = str(fila[columna['POLIZAS_CAMPANA']].value)

                fechaCreacionUnida = fechaUnida(fila[columna['FECHA_CREACION']])
                saliente = formatearSaliente(salienteEntrada)
                pk = '{0}_{1}_{2}_{3}'.format(fechaCreacionUnida, numeroPoliza, saliente, idEmpleado)
                pkClienteContacto = '{0}_{1}'.format(campanaId, salienteEntrada)

                fechaCreacion = setearFechaCelda(fila[columna['FECHA_CREACION']])
                fechaCierre = setearFechaCelda(fila[columna['FECHA_CIERRE']])
                fechaUltimaActividad = setearFechaCelda(fila[columna['ULTIMA_ACTIVIDAD']])

                if salienteEntrada is None:
                    celdaCoordenada = setearCelda2(fila[0:columna['LLAMADA_SALIENTE']+1], len(fila[0:columna['LLAMADA_SALIENTE']])-1, i)
                    LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'LLAMADA_SALIENTE': '%s;Campaña InBound/OutBound NULL;%s' % (celdaCoordenada, numeroPoliza)})
                    continue

                if fila[columna['NRO_POLIZA']].value is None or str(fila[columna['NRO_POLIZA']].value) == '':
                    celdaCoordenada = setearCelda2(fila[0:columna['NRO_POLIZA']+1], len(fila[0:columna['NRO_POLIZA']])-1, i)
                    LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'POLIZA_NULO': '%s;Numero de poliza NULL;%s' % (celdaCoordenada, numeroPoliza)})
                    continue

                if type(fechaCreacion) is not datetime.date:
                    valorErroneo = fila[columna['FECHA_CREACION']].value
                    celdaCoordenada = setearCelda2(fila[0:columna['FECHA_CREACION']+1], len(fila[0:columna['FECHA_CREACION']])-1, i)
                    mensaje = '%s;FECHA_CREACION no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'FECHA_CREACION': mensaje})
                    continue

                if type(fechaCierre) is not datetime.date:
                    fechaCierre = None
                    
                if type(fechaUltimaActividad) is not datetime.date:
                    fechaUltimaActividad = None

                if not ejecutivosExistentesDb.get(idEmpleado):
                    valorErroneo = fila[columna['ID_EMPLEADO']].value
                    celdaCoordenada = setearCelda2(fila[0:columna['ID_EMPLEADO']+1], len(fila[0:columna['ID_EMPLEADO']])-1, i)
                    mensaje = '%s;ID_EMPLEADO no existe en la DB;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'ID_EMPLEADO': mensaje})
                    continue

                nombreCampana = campanaDescripcion.get(saliente)
                exitoDuplicadoPoliza = 0
                if saliente == inbound and fechaCreacion >= fechaIncioMes and fechaCreacion <= fechaFinMes or saliente == outbound and fechaCreacion >= fechaInicioPeriodo and fechaCreacion <= fechaFinPeriodo:

                    estadoValidoReact = validarEstadoReact(estadoRetencion, estado)
                    contactoReact, nivelContactoCampana = validarContactoReact(saliente, estadoRetencion, estado, estadoUt, listaEstadoContactado)

                    if type(contactoReact) is not int:
                        valorErroneo = str(fila[columna['ESTADO_ULTIMA_TAREA']].value)
                        celdaCoordenada = setearCelda2(fila[0:columna['ESTADO_ULTIMA_TAREA']+1], len(fila[0:columna['ESTADO_ULTIMA_TAREA']])-1, i)
                        mensaje = '%s;No existe EstadoUT;%s' % (celdaCoordenada, valorErroneo)
                        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'ERROR_ESTADO_UT': mensaje})
                        continue

                    if not gestionReactTxt.get(pk):
                        gestionReactTxt[pk] = {'ESTADO_VALIDO_REACT': estadoValidoReact, 'CONTACTO_REACT': contactoReact, 'EXITO_REPETIDO_REACT': 0, 'ID_EMPLEADO': idEmpleado, 'ID_CAMPANA': campanaId, 'CAMPANA': nombreCampana, 'POLIZA': numeroPoliza, 'REPETICIONES': 1,'FECHA_CREACION': fechaCreacion, 'FECHA_CIERRE': fechaCierre, 'IN_OUT': saliente}
                        cantidadCampanasValidas += 1
                    else:

                        controlCambioPk, indiceCambio = aprobarActualizarRegistro(estado, estadoValidoReact, contactoReact, gestionReactTxt[pk]['ESTADO_VALIDO_REACT'], gestionReactTxt[pk]['CONTACTO_REACT'])
                        celdaCoordenada = setearCelda2(fila[0:columna['ID_EMPLEADO']+1], len(fila[0:columna['ID_EMPLEADO']])-1, i)
                        if controlCambioPk:
                            datosActualizados = {'ESTADO_VALIDO_REACT': estadoValidoReact, 'CONTACTO_REACT': contactoReact, 'EXITO_REPETIDO_REACT': 0, 'ID_EMPLEADO': idEmpleado, 'ID_CAMPANA': campanaId, 'CAMPANA': nombreCampana, 'POLIZA': numeroPoliza, 'REPETICIONES': gestionReactTxt[pk]['REPETICIONES'] + 1, 'FECHA_CREACION': fechaCreacion, 'FECHA_CIERRE': fechaCierre, 'IN_OUT': saliente}
                            gestionReactTxt[pk].update(datosActualizados)
                            polizaClienteDiferente = '{0}_{1}'.format(dataSalidaXlsx[pk]['CAMPANA_ID'], dataSalidaXlsx[pk]['IN_OUT'])
                            dataSalidaXlsx.pop(pk)
                            if clienteDuplicadoContactado.get(pkClienteContacto):
                                clienteDuplicadoContactado.pop(pkClienteContacto)
                            else:
                                clienteDuplicadoContactado.pop(polizaClienteDiferente)
                                LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'POLIZA_CLIENTES_DIFERENTES': '{0};POLIZA_CLIENTES_DIFERENTES;ESTADO_ANTERIOR:({1}):NUEVO_VALOR:({2}'.format(celdaCoordenada, polizaClienteDiferente, pkClienteContacto)})
                            mensaje = '{0};POLIZA_DUPLICADA;ESTADO_ANTERIOR:({1},{2}):NUEVO_VALOR:({3},{4},{5})_{6}'.format(celdaCoordenada, listaEstadoRetencionTexto.get(gestionReactTxt[pk]['ESTADO_VALIDO_REACT']), gestionReactTxt[pk]['CONTACTO_REACT'], listaEstadoRetencionTexto.get(estadoValidoReact), contactoReact, estadoRetencion, indiceCambio)
                            LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'REGISTRO_DUPLICADA': mensaje})
                        else:
                            mensaje = '{0};POLIZA_DUPLICADA;ELIMINADO:({1},{2},{3}):PERMANECE:({4},{5})'.format(celdaCoordenada, listaEstadoRetencionTexto.get(estadoValidoReact), contactoReact, estadoRetencion, listaEstadoRetencionTexto.get(gestionReactTxt[pk]['ESTADO_VALIDO_REACT']), gestionReactTxt[pk]['CONTACTO_REACT'])
                            LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'REGISTRO_DUPLICADA': mensaje})
                            gestionReactTxt[pk]['REPETICIONES'] += 1
                            continue

                    ValidacionCertificacion = 0
                    grabCertificadaReact = 0
                    if estadoValidoReact == 1:

                        ValidacionCertificacion = 2
                        if complementoCliente.get(int(numeroPoliza)) and complementoCliente[int(numeroPoliza)]['ESTADO_POLIZA'] == 'Vigente':
                            polizaReactTxt[numeroPoliza] = {'ESTADO_POLIZA_REACT': 1, 'NUMERO_POLIZA': numeroPoliza}

                        pkBaseCertificacion = '{0}_{1}'.format(str(numeroPoliza), idEmpleado)
                        if baseCertificacion.get(pkBaseCertificacion):

                            ejecutivoBaseCertificacion = baseCertificacion[pkBaseCertificacion]['ID_EMPLEADO']
                            if not ejecutivosExistentesDb.get(ejecutivoBaseCertificacion):
                                valorErroneo = baseCertificacion[pkBaseCertificacion]['ID_EMPLEADO']
                                celdaCoordenada = setearCelda2(fila[0:columna['ID_EMPLEADO']+1], len(fila[0:columna['ID_EMPLEADO']])-1, i)
                                mensaje = '%s;EJECUTIVO_BASE_CERIFICACION no existe en la DB;%s' % (celdaCoordenada, valorErroneo)
                                LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'EJECUTIVO_BASE_CERTIFICADO': mensaje})
                                continue

                            fechaLlamado = baseCertificacion[pkBaseCertificacion]['FECHA_LLAMADO']


                            if saliente == inbound and idEmpleado == ejecutivoBaseCertificacion:
                                if fechaLlamado >= fechaIncioMes and fechaLlamado <= fechaFinMes:
                                    grabCertificadaReact = 1
                                    gestionReactTxt[pk]['EXITO_REPETIDO_REACT'] = 1
                            elif saliente == outbound and idEmpleado == ejecutivoBaseCertificacion:
                                if fechaLlamado >= fechaInicioPeriodo and fechaLlamado <= fechaFinMes:
                                    grabCertificadaReact = 1
                                    gestionReactTxt[pk]['EXITO_REPETIDO_REACT'] = 1

                        certificacionReactTxt[numeroPoliza] = {'GRAB_CERTIFICADA_REACT': grabCertificadaReact, 'ID_EMPLEADO': idEmpleado, 'CANPANA': nombreCampana, 'POLIZA': numeroPoliza}

                    if grabCertificadaReact == 1:

                        ValidacionCertificacion = 1
                        if polizaExitoRepetido.get(numeroPoliza):
                            pkDataGestion = exitoRepetidoPk(numeroPoliza, polizaExitoRepetido, gestionReactTxt)

                            if pkDataGestion == 0:
                                continue

                            exitoDuplicadoPoliza = 1
                            
                            if saliente == inbound and gestionReactTxt[pkDataGestion]['IN_OUT'] == outbound:
                                if type(gestionReactTxt[pkDataGestion]['FECHA_CIERRE']) is datetime.date:
                                    if fechaCreacion >= gestionReactTxt[pkDataGestion]['FECHA_CIERRE']:
                                        gestionReactTxt[pkDataGestion]['EXITO_REPETIDO_REACT'] = 0
                                        campanasPorEjecutivos[idEmpleado][pkDataGestion]['EXITO_REPETIDO'] = 0
                                        dataSalidaXlsx[pkDataGestion]['EXITO_REPETIDO'] = exitoRepetido.get(1)
                                        dataSalidaXlsx[pkDataGestion]['ESTADO_FINAL'] = estadoFinal.get(0)
                                        exitoDuplicadoPoliza = 0
                                    else:
                                        gestionReactTxt[pk]['EXITO_REPETIDO_REACT'] = 0
                                        
                            elif saliente == outbound and gestionReactTxt[pkDataGestion]['IN_OUT'] == inbound:
                                if type(fechaCierre) is datetime.date:
                                    if fechaCierre >= gestionReactTxt[pkDataGestion]['FECHA_CREACION']:
                                        gestionReactTxt[pkDataGestion]['EXITO_REPETIDO_REACT'] = 0
                                        campanasPorEjecutivos[idEmpleado][pkDataGestion]['EXITO_REPETIDO'] = 0
                                        dataSalidaXlsx[pkDataGestion]['EXITO_REPETIDO'] = exitoRepetido.get(1)
                                        dataSalidaXlsx[pkDataGestion]['ESTADO_FINAL'] = estadoFinal.get(0)
                                        exitoDuplicadoPoliza = 0
                                    else:
                                        gestionReactTxt[pk]['EXITO_REPETIDO_REACT'] = 0
                                        
                            else:
                                if saliente == inbound:
                                    if fechaCreacion >= gestionReactTxt[pkDataGestion]['FECHA_CREACION']:
                                        gestionReactTxt[pkDataGestion]['EXITO_REPETIDO_REACT'] = 0
                                        campanasPorEjecutivos[idEmpleado][pkDataGestion]['EXITO_REPETIDO'] = 0
                                        dataSalidaXlsx[pkDataGestion]['EXITO_REPETIDO'] = exitoRepetido.get(1)
                                        dataSalidaXlsx[pkDataGestion]['ESTADO_FINAL'] = estadoFinal.get(0)
                                        exitoDuplicadoPoliza = 0
                                    else:
                                        gestionReactTxt[pk]['EXITO_REPETIDO_REACT'] = 0


                                elif saliente == outbound:
                                    if type(fechaCierre) is datetime.date and type(gestionReactTxt[pkDataGestion]['FECHA_CIERRE']) is datetime.date:
                                        if fechaCierre >= gestionReactTxt[pkDataGestion]['FECHA_CIERRE']:
                                            gestionReactTxt[pkDataGestion]['EXITO_REPETIDO_REACT'] = 0
                                            campanasPorEjecutivos[idEmpleado][pkDataGestion]['EXITO_REPETIDO'] = 0
                                            dataSalidaXlsx[pkDataGestion]['EXITO_REPETIDO'] = exitoRepetido.get(1)
                                            dataSalidaXlsx[pkDataGestion]['ESTADO_FINAL'] = estadoFinal.get(0)
                                            exitoDuplicadoPoliza = 0
                                        else:
                                            gestionReactTxt[pk]['EXITO_REPETIDO_REACT'] = 0
                                    else:
                                        valorErroneo =  '%s-VS-%s' % (fechaCierre, gestionReactTxt[pkDataGestion]['FECHA_CIERRE'])
                                        celdaCoordenada = setearCelda2(fila[0:columna['FECHA_CIERRE']+1], len(fila[0:columna['FECHA_CIERRE']])-1, i)
                                        mensaje = '%s;FECHA_CIERRE no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                                        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'FECHA_CIERRE': mensaje})
                            
                            if not polizaExitoRepetido[numeroPoliza].get(pk):
                                polizaExitoRepetido[numeroPoliza].setdefault(pk, pk)
                        else:
                            polizaExitoRepetido[numeroPoliza] = {pk: pk}
                  
                    estadoPoliza = complementoCliente[int(numeroPoliza)]['ESTADO_POLIZA']
                    estadoFinalDb = 0
                    if grabCertificadaReact == 1 and polizaReactTxt.get(numeroPoliza):
                        estadoFinalDb = 1
                        if exitoDuplicadoPoliza == 1:
                            estadoFinalDb = 0
                            
                            
                    contactoCliente, clienteUnico = validarClienteContactoDuplicado(contactoReact, nivelContactoCampana, pkClienteContacto, pk)
                    dataSalidaXlsx[pk] = {'FECHA_CREACION': fechaCreacion, 'CAMPANA': campanaEntrada, 'ESTADO': estado, 'POLIZAS_CAMPANA': polizasCampana, 'FECHA_CIERRE': fechaCierre, 'POLIZA': str(numeroPoliza), 'ESTADO_RETENCION': estadoRetencion, 'ULTIMA_ACTIVIDAD': fechaUltimaActividad, 'CAMPANA_ID': campanaId, 'ESTAD0_UT': estadoUt, 'CONTACTO_CLIENTE': contactoCliente, 'CLIENTE_UNICO': clienteUnico, 'IN_OUT': salienteEntrada, 'ID_EMPLEADO': idEmpleado, 'VALIDACION_CERTIFICACION': certificacion.get(ValidacionCertificacion), 'EXITO_REPETIDO': exitoRepetido.get(exitoDuplicadoPoliza), 'ESTADO_POLIZA': estadoPoliza, 'ESTADO_FINAL': estadoFinal.get(estadoFinalDb)}
                    
                    valoresPoliza = {'ID_EMPLEADO': idEmpleado, 'NUMERO_POLIZA': numeroPoliza, 'ESTADO_RETENCION': estadoRetencion, 'ESTAD0_UT': estadoUt, 'IN_OUT': nombreCampana, 'VALIDACION_CERTIFICACION': ValidacionCertificacion, 'EXITO_REPETIDO': exitoDuplicadoPoliza, 'ESTADO_POLIZA': estadoPoliza, 'ESTADO_FINAL': estadoFinalDb}
                    agregarCampanasPorEjecutivo(idEmpleado, pk, valoresPoliza)
                    

        if insertarPeriodoCampanaEjecutivos(campanasPorEjecutivos, fechaIncioMes):
            if insertarCampanaEjecutivos(campanasPorEjecutivos, fechaIncioMes):
                mensaje = 'InsertarCampanaEjecutivos;Se insertaron correctamente: %s Campaña(s)' % (cantidadCampanasValidas)
                LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'INSERTAR_CAMPANAS_EJECUTIVOS': mensaje})
                LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'INSERTAR_CAMPANAS_EJECUTIVOS': '-----------------------------------------------------' })

        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'FIN_CELDAS_REACTIVA': 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivoEntrada, len(tuple(hoja.rows)))})
        LOG_PROCESO_REACTIVA.setdefault(len(LOG_PROCESO_REACTIVA)+1, {'PROCESO_REACTIVA': 'Proceso del Archivo: %s Finalizado' % archivoEntrada})
        salidaGestionReactTxt = convertirDataReact(gestionReactTxt)
        dataSalida = [
            {'NOMBRE_ARCHIVO': archivoSalida['GESTION']['NOMBRE_SALIDA'], 'DATA': salidaGestionReactTxt, 'ENCABEZADO': archivoSalida['GESTION']['ENCABEZADO']},
            {'NOMBRE_ARCHIVO': archivoSalida['POLIZA']['NOMBRE_SALIDA'], 'DATA': polizaReactTxt, 'ENCABEZADO': archivoSalida['POLIZA']['ENCABEZADO']},
            {'NOMBRE_ARCHIVO': archivoSalida['CERTIFICACION']['NOMBRE_SALIDA'], 'DATA': certificacionReactTxt, 'ENCABEZADO': archivoSalida['CERTIFICACION']['ENCABEZADO']}
        ]
        return dataSalida, dataSalidaXlsx

    except Exception as e:
        LOG_PROCESO_REACTIVA.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_REACTIVA)+1: traceback.format_exc()})
        LOG_PROCESO_REACTIVA.setdefault('PROCESO_PROACTIVA', {len(LOG_PROCESO_REACTIVA)+1: 'Error al procesar Archivo: %s' % archivoEntrada})
        return False, False

# uno = '202205'
# dos = '20220501'
# tres = '20220531'
# x = r'REACTIVA\INPUTS\202205_Gestion_CoRet_Reactiva.xlsx'
# y = r'REACTIVA\INPUTS\202205_Base_Certificacion_Reactiva.xlsx'
# z = r'REACTIVA\INPUTS\202205_Complemento_Cliente_Coret.xlsx'
# rep = r'REACTIVA\OUTPUTS'
# print(leerArchivoReactiva(x, uno, dos, tres, y, z))