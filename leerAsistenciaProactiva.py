import unicodedata

from openpyxl import load_workbook
from tqdm import tqdm

from conexio_db import conectorDB
from config_xlsx import ASISTENCIA_CONFIG_XLSX
from diccionariosDB import buscarRutEjecutivosDb
from validaciones_texto import (convertirALista, formatearRut, primerDiaMes,
                                setearCelda, setearCelda2, ultimoDiaMes,
                                validarEncabezadoXlsx)

LOG_PROCESO_ASISTENCIA = dict()


def insertarEjecutivo(idEjecutivo, plataforma):
    try:
        db = conectorDB()
        cursor = db.cursor()
        sql = """MERGE ejecutivos_2 AS target
                USING (VALUES (?)) AS source (id_empleado)
                ON (source.id_empleado = target.id_empleado)
                WHEN MATCHED
                THEN UPDATE
                    SET target.plataforma = ?
                WHEN NOT MATCHED
                THEN INSERT (id_empleado, plataforma)
                    VALUES (?, ?);"""
        valores = (idEjecutivo, plataforma, idEjecutivo, plataforma)
        cursor.execute(sql, valores)
        db.commit()
        return True
    except Exception as e:
        raise Exception('Error al insertar ejecutivo: %s' % (e))
    finally:
        cursor.close()
        db.close()

def leerArchivoAsistencia(archivo, periodo):
    try:
        xls = load_workbook(archivo, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        correlativo = 1
        filaSalidaXls = dict()
        fechaProceso = primerDiaMes(periodo)

        for fila in tqdm(iterable=hoja.iter_rows(min_row=3, min_col=1), total = len(tuple(hoja.rows)), desc='Leyendo AsistenciaCRO' , unit=' Fila'):
        # for fila in hoja.rows:
            if fila[0].value is not None and fila[1].value is not None:
    

                idEjecutivo = fila[0].value
                plataforma = str(fila[1].value)

                if not filaSalidaXls.get(idEjecutivo):
                    filaSalidaXls[idEjecutivo] = {'ID_EJECUTIVO': idEjecutivo, 'PLATAFORMA': plataforma}
                    insertarEjecutivo(idEjecutivo, plataforma)
                    correlativo += 1
                else:
                    print('Ejecutivo duplicado: {idejecutivo}'.format(idejecutivo = idEjecutivo))
        return filaSalidaXls
    except Exception as e:
        print('Error no manejado: {error}'.format(error = e))
        return False, False

leerArchivoAsistencia('PROACTIVA/INPUTS/202101_Asistencia Plataforma.xlsx', '202003')
