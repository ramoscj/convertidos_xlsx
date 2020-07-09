from openpyxl import load_workbook
from tqdm import tqdm

def archivo_fuga(archivo, periodo):
    try:
        data_txt = []
        encabezado = ['LPATTR_PER_RES', 'LLAVEA', 'LPATTR_COD_POLI', 'LPATTR_COD_ORIGEN', 'TIPO', 'LPATTR_COD_STAT', 'NRO_POLIZA', 'ESTADOPOLIZA', 'FECHAINICIOVIGENCIA', 'RUT_CRO', 'NOMBRE_CRO', 'FECHAPROCESO', 'CONSIDERAR_FUGA']
        encabezado_txt = ['CRR', 'FUGA', 'STOCK', 'RUT']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        i = 0
        archivo_correcto = False
        for fila in hoja['A1:M1']:
            for celda in fila:
                if str(celda.value).upper() == encabezado[i]:
                    archivo_correcto = True
                else:
                    print('Error columna: %s' % i)
                    archivo_correcto = False
                i += 1
        if archivo_correcto:
            i = 1
            for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo DATA' , unit='Row'):
                data_xls = []
                if periodo == str(fila[0].value) and fila[9].value is not None:
                    data_xls.append(i)
                    if str(fila[4].value).upper() == 'FUGA' and str(fila[12].value).upper() != 'NO':
                        data_xls.append(1)
                    else:
                        data_xls.append(0)
                    if str(fila[5].value).upper() != 'NVIG' or str(fila[12].value).upper() == 'NO':
                        data_xls.append(1)
                    else:
                        data_xls.append(0)
                    numero_1, separador, numero_2 = str(fila[9].value).partition("-")
                    data_xls.append('%s%s' % (numero_1, numero_2))
                if len(data_xls) > 0:
                    data_txt.append(data_xls)
                    i += 1
            return data_txt, encabezado_txt
        else:
            print('Error el archivo presenta incosistencias en el encabezado')
            return False, False
    except Exception as e:
        print('Error al leer archivo: %s | %s' % (archivo, e))