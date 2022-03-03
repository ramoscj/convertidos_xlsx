import datetime

from openpyxl import load_workbook
from tqdm import tqdm

from config_xlsx import CODM_XLSX
from diccionariosDB import buscarRutEjecutivosDb, listaEstadoUtCro, listaEstadoUtContactoCro
from validaciones_texto import (primerDiaMes, setearCelda, setearFechaCelda,
                                setearFechaInput, ultimoDiaMes, primerDiaMes,
                                setearCelda2)

LOG_PROCESO_CODM = dict()

def getEstado(celdaFila):
    listaEstado = {'Pendiente': 1, 'Terminado con Exito': 2, 'Terminado sin Exito': 3}
    if listaEstado.get(celdaFila.value):
        return listaEstado[celdaFila.value]
    elif celdaFila.value == 'Sin Gestion':
        return 0
    else:
        celdaCoordenada = setearCelda(celdaFila)
        error = 'Celda%s;No existe estado;%s' % (celdaCoordenada, celdaFila.value)
        return error

def getEstadoUt(celdaFila, listaEstadoUt):
    estadoUt = str(celdaFila.value).upper()
    if listaEstadoUt.get(estadoUt):
        return listaEstadoUt[estadoUt]
    elif celdaFila.value is None or estadoUt == '':
        return 0
    else:
        celdaCoordenada = setearCelda(celdaFila)
        error = 'Celda%s;No existe estadoUt;%s' % (celdaCoordenada, celdaFila.value)
        return error

def definirEstadoValido(estado, estadoUt):
    noGestionable = {9: 'Número equivocado', 10: 'Numero invalido', 11: 'Solicita renuncia', 13: 'Cliente desconoce venta', 16: 'Sin telefono registrado'}
    if noGestionable.get(estadoUt):
        estadoValido = 'No gestionable'
    else:
        estadoValido = estado
    return estadoValido

def leerArchivoCoDm(archivoEntrada, periodo, fechaInicioEntrada, fechaFinEntrada):
    try:
        LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'INICIO_LECTURA_GESTION': 'Iniciando proceso de lectura del Archivo: %s' % archivoEntrada})
        encabezadoTxt = CODM_XLSX['ENCABEZADO_TXT']
        columna = CODM_XLSX['COLUMNAS_PROCESO_XLSX']
        xls = load_workbook(archivoEntrada, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]


        LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'ENCABEZADO_GESTION': 'Encabezado del Archivo: %s OK' % archivoEntrada})
        filaSalidaTxt = dict()
        filaSalidaXlsx = dict()
        ejecutivosExistentesDb = buscarRutEjecutivosDb(ultimoDiaMes(periodo), primerDiaMes(periodo))
        listaEstadoUt = listaEstadoUtCro()
        listaEstadoContactado = listaEstadoUtContactoCro()

        fechaInicioPeriodo = setearFechaInput(fechaInicioEntrada)
        fechaFinPeriodo = setearFechaInput(fechaFinEntrada)
        i = 0
        correlativo = 1
        LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'INICIO_CELDAS_GESTION': 'Iniciando lectura de Celdas del Archivo: %s' % archivoEntrada})

        for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo CODM' , unit=' Fila'):

            if i >= 1:

                fechaCreacion = setearFechaCelda(fila[columna['FECHA_DE_CREACION']])
                estado = getEstado(fila[columna['ESTADO']])
                nombreCampana = str(fila[columna['NOMBRE_DE_CAMPAÑA']].value)
                campanhaId = str(fila[columna['CAMPAÑA_ID']].value)
                estadoUt = getEstadoUt(fila[columna['ESTADO_UT']], listaEstadoUt)
                idEmpleado = str(fila[columna['ID_EMPLEADO']].value)

                if type(fechaCreacion) is not datetime.date:
                    valorErroneo = str(fila[columna['FECHA_DE_CREACION']].value)
                    celdaCoordenada = setearCelda2(fila[0:columna['FECHA_DE_CREACION']+1], len(fila[0:columna['FECHA_DE_CREACION']])-1, i)
                    mensaje = '%s;FECHA_DE_CREACION no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'FECHA_DE_CREACION': mensaje})
                    continue

                if type(estado) is not int:
                    valorErroneo = str(fila[columna['ESTADO']].value)
                    celdaCoordenada = setearCelda2(fila[0:columna['ESTADO']+1], len(fila[0:columna['ESTADO']])-1, i)
                    mensaje = '%s;ESTADO no existe;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'ERROR_ESTADO': mensaje})
                    continue

                if type(estadoUt) is not int:
                    valorErroneo = str(fila[columna['ESTADO_UT']].value)
                    celdaCoordenada = setearCelda2(fila[0:columna['ESTADO_UT']+1], len(fila[0:columna['ESTADO_UT']])-1, i)
                    mensaje = '%s;ESTADO_UT no existe;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'ERROR_ESTADOUT': estadoUt})
                    continue
                
                if ejecutivosExistentesDb.get(idEmpleado):

                    if fechaCreacion >= fechaInicioPeriodo and fechaCreacion <= fechaFinPeriodo:
                        filaSalidaTxt[correlativo] = {'CRR': correlativo, 'ESTADO': estado, 'ESTADO_UT': estadoUt, 'ID_CAMPANHA': campanhaId, 'CAMPANA': nombreCampana[0:30], 'ID_EMPLEADO': idEmpleado}
                        
                        contacto = 'NO'
                        if listaEstadoContactado.get(str(fila[columna['ESTADO_UT']].value).upper()) or estado == 2:
                            contacto = 'SI'
                            
                        fechaCierre = setearFechaCelda(fila[columna['FECHA_CIERRE']])
                        if type(fechaCierre) is not datetime.date:
                            fechaCierre = None
                            
                        filaSalidaXlsx[correlativo] = {'CAMPANA_ID': campanhaId, 'FECHA_CREACION': fechaCreacion, 'NOMBRE_CAMPANA': nombreCampana, 'ESTADO_UT': fila[columna['ESTADO_UT']].value, 'ESTADO': fila[columna['ESTADO']].value, 'FECHA_CIERRE': fechaCierre, 'ID_EMPLEADO': idEmpleado, 'ESTADO_VALIDO': definirEstadoValido(fila[columna['ESTADO']].value, estadoUt), 'CONTACTO': contacto}
                        correlativo += 1

                else:
                    errorRut = 'Celda%s;No existe Ejecutivo;%s' % (setearCelda(fila[columna['CAMPAÑA_ID']]), idEmpleado)
                    LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'EJECUTIVO_NO_EXISTE': errorRut})
            i += 1
        LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'FIN_CELDAS_GESTION': 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivoEntrada, len(tuple(hoja.rows)))})
        LOG_PROCESO_CODM.setdefault(len(LOG_PROCESO_CODM)+1, {'PROCESO_GESTION': 'Proceso del Archivo: %s Finalizado' % archivoEntrada})

        return filaSalidaTxt, encabezadoTxt, filaSalidaXlsx

    except Exception as e:
        errorMsg = 'Error: %s | %s' % (archivoEntrada, e)
        LOG_PROCESO_CODM.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_CODM)+1: errorMsg})
        LOG_PROCESO_CODM.setdefault('PROCESO_GESTION', {len(LOG_PROCESO_CODM)+1: 'Error al procesar Archivo: %s' % archivoEntrada})
        return False, False