from openpyxl import load_workbook
from tqdm import tqdm

from config_xlsx import CAMPANHAS_CONFIG_XLSX
from diccionariosDB import buscarRutEjecutivosDb
from validaciones_texto import (formatearRut, setearCelda2,
                                validarEncabezadoXlsx, ultimoDiaMes, primerDiaMes)

LOG_PROCESO_CAMPANHAS = dict()

def leerArchivoCampanhasEsp(archivo, periodo):
    try:
        LOG_PROCESO_CAMPANHAS.setdefault('INICIO_LECTURA_CAMPANHAS', {len(LOG_PROCESO_CAMPANHAS)+1: 'Iniciando proceso de lectura del Archivo: %s' % archivo})
        encabezadoXls = CAMPANHAS_CONFIG_XLSX['ENCABEZADO_XLSX']
        encabezadoTxt = CAMPANHAS_CONFIG_XLSX['ENCABEZADO_TXT']
        celda = CAMPANHAS_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        coordenadaEcabezado = CAMPANHAS_CONFIG_XLSX['COORDENADA_ENCABEZADO']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        archivo_correcto = validarEncabezadoXlsx(hoja[coordenadaEcabezado], encabezadoXls, archivo)
        LOG_PROCESO_CAMPANHAS.setdefault('ENCABEZADO_CAMPANHAS', {len(LOG_PROCESO_CAMPANHAS)+1: 'Encabezado del Archivo: %s OK' % archivo})
        if type(archivo_correcto) is not dict:
            i = 0
            correlativo = 1
            filaSalidaXls = dict()
            ejecutivosExistentesDb = buscarRutEjecutivosDb(ultimoDiaMes(periodo), primerDiaMes(periodo))
            for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo CamapañasEspeciales' , unit=' Fila'):

                if i >= 2 and fila[celda['ID_EMPLEADO']].value is not None:
                    idEmpleado = str(fila[celda['ID_EMPLEADO']].value)
                    if ejecutivosExistentesDb.get(idEmpleado):
                        numeroGestiones = fila[celda['NUMERO_GESTIONES']].value
                        filaSalidaXls[idEmpleado] = {'CRR': correlativo, 'NUMERO_GESTIONES': numeroGestiones, 'ID_EMPLEADO': idEmpleado}
                        correlativo += 1
                    else:
                        errorRut = '%s;No existe Ejecutivo;%s' % (setearCelda2(fila[celda['ID_EMPLEADO']], 0), idEmpleado)
                        LOG_PROCESO_CAMPANHAS.setdefault(len(LOG_PROCESO_CAMPANHAS)+1, {'EJECUTIVO_NO_EXISTE_%s' % i: errorRut})
                i += 1

            LOG_PROCESO_CAMPANHAS.setdefault(len(LOG_PROCESO_CAMPANHAS)+1, {'FIN_CELDAS_CAMPANHAS': 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivo, len(tuple(hoja.rows)))})
            LOG_PROCESO_CAMPANHAS.setdefault(len(LOG_PROCESO_CAMPANHAS)+1, {'PROCESO_CAMPANHAS': 'Proceso del Archivo: %s Finalizado' % archivo})
            return filaSalidaXls, encabezadoTxt
        else:
            LOG_PROCESO_CAMPANHAS.setdefault('ENCABEZADO_CAMPANHAS', archivo_correcto)
            raise
    except Exception as e:
        errorMsg = 'Error: %s | %s' % (archivo, e)
        LOG_PROCESO_CAMPANHAS.setdefault(len(LOG_PROCESO_CAMPANHAS)+1, {'LECTURA_ARCHIVO': errorMsg})
        LOG_PROCESO_CAMPANHAS.setdefault(len(LOG_PROCESO_CAMPANHAS)+1, {'PROCESO_CAMPANHAS': 'Error al procesar Archivo: %s' % archivo})
        return False, False
