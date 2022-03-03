import datetime

from openpyxl import load_workbook
from tqdm import tqdm

from conexio_db import conectorDB
from config_xlsx import GESTION_CONFIG_XLSX
from diccionariosDB import buscarCamphnasDb, buscarRutEjecutivosDb, listaEstadoUtCro, listaEstadoUtContactoCro
from validaciones_texto import (primerDiaMes, setearCelda, setearFechaCelda,
                                setearFechaInput, ultimoDiaMes, primerDiaMes,
                                setearCelda2)

LOG_PROCESO_GESTION = dict()

def extraerPropietariosCro(archivoPropietariosXls, periodo):
    archivo = archivoPropietariosXls
    LOG_PROCESO_GESTION.setdefault('INICIO_LECTURA_PROPIETARIOS', {len(LOG_PROCESO_GESTION)+1: 'Iniciando proceso de lectura del Archivo: %s' % archivo})
    try:
        encabezadoXls = GESTION_CONFIG_XLSX['ENCABEZADO_PROPIETARIOS_XLSX']
        celda = GESTION_CONFIG_XLSX['COLUMNAS_PROPIETARIOS_XLSX']
        coordenadaEcabezado = GESTION_CONFIG_XLSX['COORDENADA_ENCABEZADO_PROPIETARIO']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        propietariosCro = dict()
        # validarArchivo = validarEncabezadoXlsx(hoja[coordenadaEcabezado], encabezadoXls, archivo)
        # ejecutivosExistentesDb = buscarRutEjecutivosDb(ultimoDiaMes(periodo), primerDiaMes(periodo))
        i = 0

        # if type(validarArchivo) is not dict:
        for fila in tqdm(iterable=hoja.rows, total= len(tuple(hoja.rows)), desc= 'Leyendo PropietariosCRO' , unit=' Fila'):

            if i >= 1:
                campahnaId = str(fila[celda['CAMPAÑA_ID']].value)
                idEmpleado = str(fila[celda['ID_EMPLEADO']].value)
                fecha = None

                if fila[celda['FECHA']].value is not None or str(fila[celda['FECHA']].value) != '':
                    fecha = setearFechaCelda(fila[celda['FECHA']])

                if fila[celda['ID_EMPLEADO']].value is None or str(fila[celda['ID_EMPLEADO']].value) == '':
                    celdaCoordenada = setearCelda2(fila[0:celda['ID_EMPLEADO']+1], len(fila[0:celda['ID_EMPLEADO']])-1, i)
                    LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1 , {'EMPLEADO_PROPIETARIO_NULL': '{0};El ID_EMPLEADO es NULL;{1}'.format(celdaCoordenada, idEmpleado)})
                    continue

                if not propietariosCro.get(campahnaId):
                    propietariosCro[campahnaId] = {'ID_EMPLEADO': idEmpleado, 'FECHA': fecha}
                else:
                    if fecha is not None:
                        if propietariosCro[campahnaId]['FECHA'] is None:
                            propietariosCro[campahnaId]['ID_EMPLEADO'] = idEmpleado
                            propietariosCro[campahnaId]['FECHA'] = fecha
                        else:
                            if fecha > propietariosCro[campahnaId]['FECHA']:
                                propietariosCro[campahnaId]['ID_EMPLEADO'] = idEmpleado
                                propietariosCro[campahnaId]['FECHA'] = fecha
            i += 1

        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1 , {'ENCABEZADO_PROPIETARIOSCRO': 'Encabezado del Archivo: %s OK' % archivo})
        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'LECTURA_PROPIETARIOS': 'Lectura del Archivo: %s Finalizado' % archivo})
        return propietariosCro
        # else:
        #     LOG_PROCESO_GESTION.setdefault('ENCABEZADO_PROPIETARIOSCRO', validarArchivo)
        #     raise
    except Exception as e:
        errorMsg = 'Error al leer archivo: %s | %s' % (archivo, e)
        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1 , {'LECTURA_PROPIETARIOSCRO': errorMsg})
        raise

def getEstado(celdaFila):
    listaEstado = {'Pendiente': 1, 'Terminado con Exito': 2, 'Terminado sin Exito': 3}
    if listaEstado.get(celdaFila.value):
        return listaEstado[celdaFila.value]
    elif celdaFila.value == 'Sin Gestion':
        return 0
    else:
        celdaCoordenada = setearCelda(celdaFila)
        error = 'Celda%s;No existe estado;%s' % (celdaCoordenada, celdaFila.value)
        return error

def getEstadoUt(celdaFila, listaEstadoUt):
    estadoUt = str(celdaFila.value).upper()
    if listaEstadoUt.get(estadoUt):
        return listaEstadoUt[estadoUt]
    elif celdaFila.value is None or str(celdaFila.value) == '':
        return 0
    else:
        celdaCoordenada = setearCelda(celdaFila)
        error = 'Celda%s;No existe estadoUt;%s' % (celdaCoordenada, celdaFila.value)
        return error

def insertarCamphnaCro(nombresCampahnasNuevas):
    try:
        db = conectorDB()
        cursor = db.cursor()
        sql = """INSERT INTO codigos_cro (nombre) VALUES (?);"""
        cursor.executemany(sql, nombresCampahnasNuevas)
        db.commit()
        return True
    except Exception as e:
        raise Exception('Error al insertar nombresCampahnasNuevas: {0}'.format(e))
    finally:
        cursor.close()
        db.close()

def leerArchivoGestion(archivoEntrada, periodo, fechaInicioEntrada, fechaFinEntrada, archivoPropietariosXls):
    try:
        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'INICIO_LECTURA_GESTION': 'Iniciando proceso de lectura del Archivo: %s' % archivoEntrada})
        encabezadoTxt = GESTION_CONFIG_XLSX['ENCABEZADO_TXT']
        columna = GESTION_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        xls = load_workbook(archivoEntrada, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]


        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'ENCABEZADO_GESTION': 'Encabezado del Archivo: %s OK' % archivoEntrada})
        dataSalida = dict()
        dataSalidaXlsx = dict()
        campanasNuevas = []
        propietarioCro = extraerPropietariosCro(archivoPropietariosXls, periodo)
        campahnasExistentesDb = buscarCamphnasDb()
        ejecutivosExistentesDb = buscarRutEjecutivosDb(ultimoDiaMes(periodo), primerDiaMes(periodo))
        listaEstadoUt = listaEstadoUtCro()
        listaEstadoContactado = listaEstadoUtContactoCro()

        fechaInicioPeriodo = setearFechaInput(fechaInicioEntrada)
        fechaFinPeriodo = setearFechaInput(fechaFinEntrada)
        fechaIncioMes = primerDiaMes(periodo)
        fechaFinMes = ultimoDiaMes(periodo)
        i = 0
        correlativo = 1
        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'INICIO_CELDAS_GESTION': 'Iniciando lectura de Celdas del Archivo: %s' % archivoEntrada})

        for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo GestionCRO' , unit=' Fila'):

            if i >= 1:

                fechaCreacion = setearFechaCelda(fila[columna['FECHA_DE_CREACION']])
                fechaCierre = setearFechaCelda(fila[columna['FECHA_DE_CIERRE']])
                estado = getEstado(fila[columna['ESTADO']])
                nombreCampana = str(fila[columna['NOMBRE_DE_CAMPAÑA']].value)
                campanhaId = str(fila[columna['CAMPAÑA_ID']].value)
                estadoUt = getEstadoUt(fila[columna['ESTADO_UT']], listaEstadoUt)
                idEmpleado = str(fila[columna['ID_EMPLEADO']].value)

                if type(fechaCreacion) is not datetime.date:
                    valorErroneo = str(fila[columna['FECHA_DE_CREACION']].value)
                    celdaCoordenada = setearCelda2(fila[0:columna['FECHA_DE_CREACION']+1], len(fila[0:columna['FECHA_DE_CREACION']])-1, i)
                    mensaje = '%s;FECHA_DE_CREACION no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'FECHA_DE_CREACION': mensaje})
                    continue

                if type(estado) is not int:
                    valorErroneo = str(fila[columna['ESTADO']].value)
                    celdaCoordenada = setearCelda2(fila[0:columna['ESTADO']+1], len(fila[0:columna['ESTADO']])-1, i)
                    mensaje = '%s;ESTADO no existe;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'ERROR_ESTADO': mensaje})
                    continue

                if type(estadoUt) is not int and estado != 0:
                    valorErroneo = str(fila[columna['ESTADO_UT']].value)
                    celdaCoordenada = setearCelda2(fila[0:columna['ESTADO_UT']+1], len(fila[0:columna['ESTADO_UT']])-1, i)
                    mensaje = '%s;ESTADO_UT no existe;%s' % (celdaCoordenada, valorErroneo)
                    LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'ERROR_ESTADOUT': estadoUt})
                    raise Exception('Error en la columna ESTADO_UT')
                elif type(estadoUt) is not int:
                    estadoUt = 0

                if nombreCampana == 'Inbound CRO':
                    if estado != 0:
                        fechaUltimaModificacion = fechaCierre
                        if type(fechaUltimaModificacion) is not datetime.date:
                            errorCampana = 'Celda%s;FECHA_CIERRE No es valida;%s' % (setearCelda(fila[columna['CAMPAÑA_ID']]), fechaUltimaModificacion)
                            LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'FECHA_CIERRE_ERROR': errorCampana})
                            continue
                        if fechaUltimaModificacion >= fechaIncioMes and fechaUltimaModificacion <= fechaFinMes:
                            if propietarioCro.get(campanhaId):
                                ejecutivoCorrecto = propietarioCro[campanhaId]['ID_EMPLEADO']
                            else:
                                ejecutivoCorrecto = idEmpleado
                        else:
                            continue
                    else:
                        continue
                else:
                    if fechaCreacion < fechaInicioPeriodo or fechaCreacion > fechaFinPeriodo:
                        continue
                    if propietarioCro.get(campanhaId):
                        ejecutivoCorrecto = propietarioCro[campanhaId]['ID_EMPLEADO']
                    else:
                        ejecutivoCorrecto = idEmpleado

                if not campahnasExistentesDb.get(nombreCampana):
                    campahnasExistentesDb[nombreCampana] = {'NOMBRE_CAMPANA': nombreCampana}
                    campanasNuevas.append([nombreCampana])
                    
                if type(fechaCierre) is not datetime.date:
                    fechaCierre = None

                if ejecutivosExistentesDb.get(ejecutivoCorrecto):
                    dataSalida[correlativo] = {'CRR': correlativo, 'ESTADO': estado, 'ESTADO_UT': estadoUt, 'ID_CAMPANHA': campanhaId, 'CAMPANA': nombreCampana[0:30], 'ID_EMPLEADO': ejecutivoCorrecto}

                    contacto = 'NO'
                    if listaEstadoContactado.get(str(fila[columna['ESTADO_UT']].value).upper()) or estado == 2:
                        contacto = 'SI'
                        
                    ejecutivoPropietario = None
                    if propietarioCro.get(campanhaId):
                        ejecutivoPropietario = propietarioCro[campanhaId]['ID_EMPLEADO']
                        
                    dataSalidaXlsx[correlativo] = {'ID_CAMPANHA': campanhaId, 'FECHA_CREACION': fechaCreacion, 'CAMPANA': nombreCampana, 'ESTADO_UT': fila[columna['ESTADO_UT']].value, 'ESTADO': fila[columna['ESTADO']].value, 'FECHA_CIERRE': fechaCierre, 'ID_EMPLEADO': idEmpleado, 'ID_EMPLEADO2': ejecutivoPropietario, 'EJECUTIVO_CORRECTO': ejecutivoCorrecto, 'CONTACTO': contacto}
                    correlativo += 1
                else:
                    errorRut = 'Celda%s;No existe Ejecutivo;%s' % (setearCelda(fila[columna['CAMPAÑA_ID']]), ejecutivoCorrecto)
                    LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'EJECUTIVO_NO_EXISTE': errorRut})
            i += 1
        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'FIN_CELDAS_GESTION': 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivoEntrada, len(tuple(hoja.rows)))})
        LOG_PROCESO_GESTION.setdefault(len(LOG_PROCESO_GESTION)+1, {'PROCESO_GESTION': 'Proceso del Archivo: %s Finalizado' % archivoEntrada})
        if len(campanasNuevas) > 0:
            insertarCamphnaCro(campanasNuevas)
        return dataSalida, encabezadoTxt, dataSalidaXlsx

    except Exception as e:
        errorMsg = 'Error: %s | %s' % (archivoEntrada, e)
        LOG_PROCESO_GESTION.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_GESTION)+1: errorMsg})
        LOG_PROCESO_GESTION.setdefault('PROCESO_GESTION', {len(LOG_PROCESO_GESTION)+1: 'Error al procesar Archivo: %s' % archivoEntrada})
        return False, False