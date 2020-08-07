from tqdm.auto import trange
from time import sleep
from tqdm import tqdm
from openpyxl import load_workbook
import os.path

# for i in trange(4, desc='1st loop'):
#     for j in trange(5, desc='2nd loop', leave=False):
#         for k in trange(50, desc='3rd loop', leave=False):
#             sleep(0.01)

from alive_progress import alive_bar
import time
filenames = ['INPUTS/202005_Asistencia_CRO.xlsx', 'INPUTS/201911_Fuga_Agencia.xlsx']
for x in range(0, len(filenames)):
    xls = load_workbook(filenames[x], read_only=True, data_only=True)
    nombre_hoja = xls.sheetnames
    hoja = xls[nombre_hoja[0]]
    with alive_bar(len(hoja.rows)) as bar:
        # for filename in tqdm(iterable= filenames, total= len(filenames), desc="Files"):
        # for row in tqdm(iterable=hoja.rows, total= len(tuple(hoja.rows)), desc= 'Leyendo PropietariosCRO' , unit=' Fila'):
        for row in hoja.rows:
            for cell in row:
                # print(cell)
                pass
            bar()

# with alive_bar(1000) as bar:
#     for item in range(1000):
#         if item % 300 == 0:
#             print('Encontrado')
#         time.sleep(.01)
#         bar()