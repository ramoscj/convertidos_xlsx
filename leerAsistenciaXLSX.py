import unicodedata
import datetime
import traceback

from openpyxl import load_workbook
from tqdm import tqdm

from conexio_db import conectorDB
from config_xlsx import ASISTENCIA_CONFIG_XLSX
from validaciones_texto import (formatearRut, primerDiaMes, setearCelda,
                                setearCelda2, ultimoDiaMes,
                                validarEncabezadoXlsx, convertirALista)
from diccionariosDB import buscarRutEjecutivosDb

LOG_PROCESO_ASISTENCIA = dict()

def updateEjecutivoFechaDesv(periodo):
    try:
        db = conectorDB()
        cursor = db.cursor()
        ultimoDia = ultimoDiaMes(periodo)
        sql = """UPDATE ejecutivos SET fecha_desvinculacion = ? WHERE fecha_desvinculacion is NULL"""
        cursor.execute(sql, (ultimoDia,))
        db.commit()
        return True
    except Exception as e:
        raise Exception('Error al actualizar tabla de ejecutivos | %s' %e)
    finally:
        cursor.close()
        db.close()

def updateEjecutivoFechaIngreso(ejecutivosExistentes):
    try:
        db = conectorDB()
        cursor = db.cursor()
        sql = """UPDATE ejecutivos SET fecha_ingreso = ? WHERE id_empleado = ? AND fecha_ingreso > ?"""
        cursor.executemany(sql, ejecutivosExistentes)
        db.commit()
        return True
    except Exception as e:
        raise Exception('Error al actualizar tabla de ejecutivos | %s' %e)
    finally:
        cursor.close()
        db.close()

def insertarEjecutivo(empleadosLista):
    try:
        db = conectorDB()
        cursor = db.cursor()
        sql = """MERGE ejecutivos AS target
                USING (VALUES (?)) AS source (id_empleado)
                ON (source.id_empleado = target.id_empleado)
                WHEN MATCHED
                THEN UPDATE
                    SET target.plataforma = ?,
                        target.fecha_desvinculacion = NULL
                WHEN NOT MATCHED
                THEN INSERT (id_empleado, plataforma, fecha_ingreso, fecha_desvinculacion)
                    VALUES (?, ?, ?, NULL);"""
        cursor.executemany(sql, empleadosLista)
        db.commit()
        return True
    except Exception as e:
        raise Exception('Error al insertarEjecutivo: {0}'.format(e))
    finally:
        cursor.close()
        db.close()

def calcularDiasHablies(columnas):
    diasDeSemana = {'LUNES': 1, 'MARTES': 2, 'MIERCOLES': 3, 'JUEVES': 4, 'VIERNES': 5}
    diasHabiles = 0
    for columna in columnas:
        for celda in columna:
            nfkd_form = unicodedata.normalize('NFKD', str(celda.value).upper())
            diaSinAcento = nfkd_form.encode('ASCII', 'ignore')
            diaFormateado = str(diaSinAcento.decode('utf-8')).strip()
            if diasDeSemana.get(diaFormateado):
                diasHabiles += 1
    return diasHabiles

def leerArchivoAsistencia(archivo, periodo):
    try:
        LOG_PROCESO_ASISTENCIA.setdefault('INICIO_LECTURA_ASISTENCIA', {len(LOG_PROCESO_ASISTENCIA)+1: 'Iniciando proceso de lectura del Archivo: %s' % archivo})
        encabezadoTxt = ASISTENCIA_CONFIG_XLSX['ENCABEZADO_TXT']
        columna = ASISTENCIA_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        columnasAdicionales = ASISTENCIA_CONFIG_XLSX['COLUMNAS_ADICIONALES']
        xls = load_workbook(archivo, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        correlativo = 1
        
        LOG_PROCESO_ASISTENCIA.setdefault(len(LOG_PROCESO_ASISTENCIA)+1, {'ENCABEZADO_ASISTENCIA': 'Encabezado del Archivo: %s OK' % archivo})
        filaSalidaXls = dict()
        empleadosLista = dict()
        ejecutivosExistentes = []
        fechaIncioMes = primerDiaMes(periodo)
        fechaFinMes = ultimoDiaMes(periodo)
        ejecutivosExistentesDb = buscarRutEjecutivosDb(fechaFinMes, fechaIncioMes)
        totalColumnas = calcularDiasHablies(hoja.iter_rows(min_row=1, min_col=3, max_row=1))
        totalFilas = len(tuple(hoja.iter_rows(min_row=3, min_col=1)))
        LOG_PROCESO_ASISTENCIA.setdefault(len(LOG_PROCESO_ASISTENCIA)+1, {'INICIO_CELDAS_ASISTENCIA': 'Iniciando lectura de Celdas del Archivo: %s' % archivo})

        updateEjecutivoFechaDesv(periodo)
        for fila in tqdm(iterable=hoja.iter_rows(min_row=3, min_col=1), total = totalFilas, desc='Leyendo AsistenciaCRO' , unit=' Fila'):
        # for fila in hoja.iter_rows(min_row=3, min_col=1):
            diasVacaciones = 0
            ausentismoMes = 1
            beneficioMes = 0
            ajusteReintegro = 0
            if fila[columna['ID_EMPLEADO']].value is not None and fila[columna['PLATAFORMA']].value is not None:

                idEmpleado = fila[columna['ID_EMPLEADO']].value
                plataforma = str(fila[columna['PLATAFORMA']].value).upper()
                fugaCartera = str(fila[columna['FUGA_CARTERA']].value).strip().upper()

                if ejecutivosExistentesDb.get(str(idEmpleado)):
                    fechaIngreso = datetime.datetime.strptime(ejecutivosExistentesDb[str(idEmpleado)]['FECHA_INGRESO'], '%d-%m-%Y')
                    if fechaIngreso.date() > fechaIncioMes:
                        ejecutivosExistentes.append([fechaIncioMes, str(idEmpleado), fechaIncioMes])

                if not empleadosLista.get(idEmpleado):
                    empleadosLista[idEmpleado] = {'ID_EMPLEADO': idEmpleado, 'PLATAFORMA': plataforma, 'ID_EMPLEADO2': idEmpleado, 'PLATAFORMA2': plataforma, 'PRIMER_DIA': fechaIncioMes}

                if not filaSalidaXls.get(idEmpleado):
                    conteoVhcAplica = 0
                    vhcAplica = 0
                    
                    for celda in range(columnasAdicionales, totalColumnas + columnasAdicionales):
                        textoAsistecia = str(fila[celda].value).upper()
                        estadoAsistencia = textoAsistecia.strip()
                        if estadoAsistencia == 'V' or estadoAsistencia == 'VAC':
                            diasVacaciones += 1
                            conteoVhcAplica += 1
                        else:
                            conteoVhcAplica = 0
                            
                        if estadoAsistencia == 'B':
                            beneficioMes += 1

                        if conteoVhcAplica == 5:
                            vhcAplica = 1
                            conteoVhcAplica = 0

                        if estadoAsistencia == '1':
                            ausentismoMes = 0

                    if fugaCartera == '1':
                        ajusteReintegro = 1
                    
                    filaSalidaXls[idEmpleado] = {'CRR': correlativo, 'VHC_MES': diasVacaciones, 'DIAS_HABILES_MES': totalColumnas, 'CARGA': vhcAplica, 'AUSENTISMO_MES': ausentismoMes, 'BENEFICIO_MES': beneficioMes, 'FUGA_CARTERA': ajusteReintegro, 'ID_EMPLEADO': idEmpleado}
                    correlativo += 1
                else:
                    errorRut = '%s - Ejecutivo duplicado: %s' % (setearCelda2(fila[columna['ID_EMPLEADO']],0), idEmpleado)
                    LOG_PROCESO_ASISTENCIA.setdefault(len(LOG_PROCESO_ASISTENCIA)+1, {'EJECUTIVO_DUPLICADO_%s' % str(len(LOG_PROCESO_ASISTENCIA)+1): errorRut})
        LOG_PROCESO_ASISTENCIA.setdefault('FIN_CELDAS_ASISTENCIA', {len(LOG_PROCESO_ASISTENCIA)+1: 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivo, len(tuple(hoja.rows)))})
        LOG_PROCESO_ASISTENCIA.setdefault('PROCESO_ASISTENCIA', {len(LOG_PROCESO_ASISTENCIA)+1: 'Proceso del Archivo: %s Finalizado' % archivo})
        if len(empleadosLista) > 0:
            listaEmpleadosFormateada = convertirALista(empleadosLista)
            insertarEjecutivo(listaEmpleadosFormateada)
        if len(ejecutivosExistentes) > 0:
            updateEjecutivoFechaIngreso(ejecutivosExistentes)
        return filaSalidaXls, encabezadoTxt

    except Exception as e:
        # errorMsg = 'Error: %s | %s' % (archivo, e)
        LOG_PROCESO_ASISTENCIA.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_ASISTENCIA)+1: traceback.format_exc()})
        LOG_PROCESO_ASISTENCIA.setdefault('PROCESO_ASISTENCIA', {len(LOG_PROCESO_ASISTENCIA)+1: 'Error al procesar Archivo: %s' % archivo})
        return False, False