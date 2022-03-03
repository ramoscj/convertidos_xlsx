import datetime
import string
import os.path

from dateutil.relativedelta import relativedelta

from openpyxl import load_workbook
import re


def validaFechaInput(fecha_x):
    try:
        f1 = fecha_x[0:4]
        f2 = fecha_x[4:6]
        if type(datetime.date(int(f1), int(f2), 1)) is datetime.date and len(fecha_x) <= 6:
            return True
        else:
            return False
    except Exception as e:
        print("Error de fecha, formato correcto YYYYMM | %s" % e)

def validaFechaCelda(celdaFila):
    try:
        fecha = str(celdaFila.value)
        fechaAnho = fecha[0:4]
        fechaMes = fecha[4:6]
        if type(datetime.date(int(fechaAnho), int(fechaMes), 1)) is datetime.date:
            return celdaFila
        else:
            errorMsg = "Celda%s - validaFechaCelda: %s | %s" % (setearCelda(celdaFila), str(celdaFila.value), e)
            return errorMsg
    except Exception as e:
        errorMsg = "Celda%s - validaFechaCelda: %s | %s" % (setearCelda(celdaFila), str(celdaFila.value), e)
        return errorMsg

def setearFechaCelda(celdaFila):
    try:
        fecha = str(celdaFila.value).replace("-","")
        fechaAnho = fecha[0:4]
        fechaMes = fecha[4:6]
        fechaDia = fecha[6:8]
        fechaSalida = datetime.date(int(fechaAnho), int(fechaMes), int(fechaDia))
        return fechaSalida
    except Exception as e:
        errorMsg = "Celda%s - setearFechaCelda: %s | %s" % (setearCelda(celdaFila), str(celdaFila.value), e)
        return errorMsg

def setearFechaInput(fecha):
    try:
        fechaAnho = str(fecha)[0:4]
        fechaMes = str(fecha)[4:6]
        fechaDia = str(fecha)[6:8]
        fechaSalida = datetime.date(int(fechaAnho), int(fechaMes), int(fechaDia))
        return fechaSalida
    except Exception as e:
        errorMsg = "Error de fecha %s, formato correcto YYYYMMDD setearFechaInput | %s" % (fecha, e)
        raise Exception(errorMsg)

def formatearRut(rut):
    rutMantisa, separador, dv = rut.partition("-")
    rutSalida = '%s%s' % (rutMantisa, dv)
    return rutSalida

def formatearRutGion(rut):
    caracteres = len(str(rut).strip())
    rutSalida = '%s-%s' % (rut[0:caracteres-1], rut[caracteres-1:caracteres])
    return rutSalida

def validarEncabezadoXlsx(filasXlsx: [], encabezadoXls: [], nombreArchivo):
    columnasError = dict()
    i = 0
    for fila in filasXlsx:
        for celda in fila:
            if str(celda.value).upper() != encabezadoXls[i]:
                celda = setearCelda(str(fila[i]))
                error = 'Celda%s <strong>%s</strong> <a style="color:red">Encabezado incorrecto:</a> %s' % (celda,  sacarNombreArchivo(nombreArchivo), encabezadoXls[i])
                columnasError.setdefault(len(columnasError)+1, error)
            i += 1
    if len(columnasError) > 0:
        return columnasError
    else:
        return True

def compruebaEncabezado(archivoXlsx, encabezadoXls, coordenadaEcabezado):
    xls = load_workbook(archivoXlsx, read_only=True)
    nombre_hoja = xls.sheetnames
    hoja = xls[nombre_hoja[0]]
    archivo_correcto = validarEncabezadoXlsx(hoja[coordenadaEcabezado], encabezadoXls, archivoXlsx)
    return archivo_correcto

def setearCelda(celda):
    resto, separador, celdaN = str(celda).partition(".")
    return ('<%s') % celdaN

def setearCelda2(celda, cantidad, *nroRegistro):
    if cantidad > 0:
        if celda[cantidad].value is None or str(celda[cantidad].value) == '':
            cantidad -= 1
            setearCelda2(celda, cantidad, nroRegistro[0])
        if cantidad == len(celda) -1:
            resto, separador, celdaN = str(celda[cantidad]).partition(".")
            coordenada = 'Celda<%s' % celdaN
        else:
            resto, separador, celdaN = str(celda[cantidad]).partition(".")
            letrasAbecedario = list(string.ascii_uppercase)
            coordenada = 'Celda<%s%s>' % (letrasAbecedario[len(celda)-1], nroRegistro[0])
    else:
        resto, separador, celdaN = str(celda).partition(".")
        coordenada = 'Celda<%s' % celdaN
    return coordenada

def primerDiaMes(fecha):
    fechaAnho = str(fecha)[0:4]
    fechaMes = str(fecha)[4:6]
    primerDia = datetime.datetime(int(fechaAnho), int(fechaMes), 1).replace(day=1).date()
    return primerDia

def ultimoDiaMes(fecha):
    fechaAnho = str(fecha)[0:4]
    fechaMes = str(fecha)[4:6]
    ultimoDia = datetime.datetime(int(fechaAnho), int(fechaMes), 1).replace(day=1).date()+relativedelta(months=1)+datetime.timedelta(days=-1)
    return ultimoDia

def formatearPlataformaCRO(plataforma):
    carateres = len(str(plataforma).strip())
    plataformaSalida = str(plataforma[0:carateres-1]).strip()
    if plataformaSalida != 'CRO':
        plataformaSalida = plataforma
    return plataformaSalida

def formatearFechaYM(fecha):
    try:
        fechaAnho = str(fecha)[0:4]
        fechaMes = str(fecha)[4:6]
        fechaSalida = datetime.date(int(fechaAnho), int(fechaMes), 1)
        return fechaSalida
    except Exception as e:
        errorMsg = "Error %s, formatearFechaYM YYYYMM | %s" % (fecha, e)
        raise Exception(errorMsg)

def formatearFechaMesSiguiente(fecha):
    try:
        fechaAnho = str(fecha)[0:4]
        fechaMes = str(fecha)[4:6]
        fechaSalida = datetime.datetime(int(fechaAnho), int(fechaMes), 1).replace(day=1).date()+relativedelta(months=1)
        return fechaSalida
    except Exception as e:
        errorMsg = "Error %s, formato correcto YYYYMM | %s" % (fecha, e)
        raise Exception(errorMsg)

def mesSiguienteUltimoDia(fecha):
    try:
        fechaAnho = str(fecha)[0:4]
        fechaMes = str(fecha)[4:6]
        fechaSalida = datetime.datetime(int(fechaAnho), int(fechaMes), 1).replace(day=1).date()+relativedelta(months=1)
        fechaSalida = datetime.datetime(int(fechaSalida.strftime("%Y")), int(fechaSalida.strftime("%m")), 1).replace(day=1).date()+relativedelta(months=1)+datetime.timedelta(days=-1)
        return fechaSalida
    except Exception as e:
        errorMsg = "Error %s, formato correcto YYYYMM | %s" % (fecha, e)
        raise Exception(errorMsg)

def separarNombreApellido(nombreCompleto):
    apellidos, separador, nombres = nombreCompleto.partition(',')
    nombresSalida = setearNombre(nombres.split())
    apellidoPaterno = apellidos.split()
    apellidoMaterno = setearApellidoMaterno(apellidos.split())
    return apellidoPaterno[0], apellidoMaterno, nombresSalida

def setearApellidoMaterno(texto):
    textoFormateado = ''
    for cantidad in range(1,len(texto)):
        if cantidad == len(texto)-1:
            textoFormateado += texto[cantidad]
        else:
            textoFormateado += texto[cantidad] + ' '
    return textoFormateado

def setearNombre(texto):
    textoFormateado = ''
    for cantidad in range(0,len(texto)):
        if cantidad == len(texto)-1:
            textoFormateado += texto[cantidad]
        else:
            textoFormateado += texto[cantidad] + ' '
    return textoFormateado

def formatearFechaMesAnterior(fecha):
    try:
        fechaAnho = str(fecha)[0:4]
        fechaMes = str(fecha)[4:6]
        fechaSalida = datetime.datetime(int(fechaAnho), int(fechaMes), 1).replace(day=1).date()-relativedelta(months=1)
        return fechaSalida
    except Exception as e:
        errorMsg = "Error %s, formato correcto YYYYMM | %s" % (fecha, e)
        raise Exception(errorMsg)

def formatearNumeroPoliza(numeroPoliza):
    polizaSalida = None
    nroCertificado = 0
    if numeroPoliza is not None and numeroPoliza != '':
        nroPolizaSolo, separador, nroCertificado = str(numeroPoliza).partition("_")
        polizaSalida = int(nroPolizaSolo)
        if nroCertificado == '':
            nroCertificado = 0
    return polizaSalida, int(nroCertificado)

def formatearIdCliente(idCliente):
    campana, separador, codCliente = str(idCliente).partition(":")
    nombreCliente, separador, rutEjecutivo = str(codCliente).partition("-")
    return nombreCliente.strip()

def convertirALista(dataLista: dict):
    listaFinal = []
    for llave, valores in dataLista.items():
        lista = []
        for valor in valores:
            lista.append(dataLista[llave][valor])
        listaFinal.append(list(lista))
    return listaFinal

def convertirListaProactiva(dataLista: dict, ejecutivosExistentes, fechaPeriodo):
    lista = []
    for idEmpleado in dataLista.keys():
        if not ejecutivosExistentes.get(str(idEmpleado)):
            lista.append([str(idEmpleado), fechaPeriodo])
    return lista

def convertirListaReactiva(dataLista: dict, ejecutivosExistentes, fechaPeriodo):
    lista = []
    for idEmpleado in dataLista.keys():
        if not ejecutivosExistentes.get(str(idEmpleado)):
            lista.append([str(idEmpleado), fechaPeriodo])
    return lista

def setearCampanasProactiva(dataLista: dict, idEjecutivo):
    lista = []
    nombreCampana = str(dataLista['NOMBRE_CAMPAÑA'])[0:30].rstrip()
    data = [idEjecutivo, dataLista['NUMERO_POLIZA'], dataLista['CAMPAÑA_ID'], nombreCampana, dataLista['ESTADO_RETENCION'], dataLista['RETENCION_COBRANZA'], dataLista['RETENCION_ACTIVACION'], dataLista['RETENCION_RL_COBRANZA'], dataLista['RETENCION_RL_ACTIVACION'],dataLista['ESTADO_VALIDO'], dataLista['ESTADO_VALIDOUT'], dataLista['FECHA_CIERRE'], dataLista['RELIQUIDACION'], dataLista['NUMERO_POLIZA_CERTIFICADO'], dataLista['POLIZAS_CAMPANA'], dataLista['NOMBRE_CAMPAÑA'], dataLista['FECHA_CREACION'], dataLista['FECHA_EXPIRACION_CORET'], dataLista['FECHA_ULTIMO_PAGO'], dataLista['FECHA_MANDATO'], dataLista['ESTADO_MANDATO']]
    lista.append(data)
    return lista

def setearCampanasReactiva(dataLista: dict, idEjecutivo):
    lista = []
    data = [idEjecutivo, dataLista['NUMERO_POLIZA'], dataLista['ESTADO_RETENCION'], dataLista['ESTAD0_UT'], dataLista['IN_OUT'], dataLista['VALIDACION_CERTIFICACION'], dataLista['EXITO_REPETIDO'], dataLista['ESTADO_POLIZA'], dataLista['ESTADO_FINAL']]
    lista.append(data)
    return lista

def convertirDataReact(dataReact: dict):
    dataFinal = dict()
    for valores in dataReact.values():
        pk = '{0}_{1}_{2}_{3}'.format(str(valores['FECHA_CREACION']), valores['POLIZA'], valores['CAMPANA'], valores['ID_EMPLEADO'])
        dataFinal[pk] = {'ESTADO_VALIDO_REACT': valores['ESTADO_VALIDO_REACT'], 'CONTACTO_REACT': valores['CONTACTO_REACT'], 'EXITO_REPETIDO_REACT': valores['EXITO_REPETIDO_REACT'], 'REPETICIONES': valores['REPETICIONES'], 'ID_EMPLEADO': valores['ID_EMPLEADO'], 'ID_CAMPANA': valores['ID_CAMPANA'], 'CAMPANA': valores['CAMPANA'], 'POLIZA': valores['POLIZA']}
    return dataFinal

def fechaMesAnterior(fecha):
    try:
        fechaSalida = fecha-relativedelta(months=1)
        return fechaSalida
    except Exception as e:
        errorMsg = "Error %s, fechaMesAnterior | %s" % (fecha, e)
        raise Exception(errorMsg)

def fechaUnida(celdaFila):
    try:
        fecha = str(celdaFila.value).replace("-","")
        fechaSalida = '{0}{1}{2}'.format(fecha[0:4], fecha[4:6], fecha[6:8])
        return fechaSalida
    except Exception as e:
        errorMsg = "Celda%s - fechaUnida: %s | %s" % (setearCelda(celdaFila), str(celdaFila.value), e)
        return errorMsg

def encontrarArchivo(archivoXlsx):
    valor = False
    if os.path.isfile(archivoXlsx):
        valor = True
    return valor

def encontrarDirectorio(pathDestino):
    valor = False
    if os.path.isdir(pathDestino):
        valor = True
    return valor

def sacarNombreArchivo(cadena):
    cadenaFormateada = str(cadena).replace("\\", "-")
    cadenaFormateada = str(cadenaFormateada).replace("/", "-")
    cadenaSalida = re.split("-", cadenaFormateada)
    return cadenaSalida[len(cadenaSalida)-1]
