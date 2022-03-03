import contextlib
import datetime

from openpyxl import load_workbook
from tqdm import tqdm

from complementoCliente import (LOG_COMPLEMENTO_CLIENTE,
                                extraerComplementoCliente)
from conexio_db import conectorDB
from config_xlsx import PROACTIVA_CONFIG_XLSX
from diccionariosDB import (CampanasPorPeriodoProactiva, buscarPolizasReliquidar,
                            buscarRutEjecutivosDb, estadoRetencionProDesc,
                            listaEstadoRetencionProactiva, listaEstadoUtAll,
                            listaEstadoUtContacto, listaEstadoUtDesc,
                            periodoCampanasProactiva)

from validaciones_texto import (convertirListaProactiva,
                                formatearFechaMesAnterior,
                                formatearNumeroPoliza, mesSiguienteUltimoDia,
                                primerDiaMes, setearCampanasProactiva,
                                setearCelda2, setearFechaCelda,
                                ultimoDiaMes, validarEncabezadoXlsx, fechaMesAnterior)

from dataXlsxProactiva import (definirEstadoRetencionPro, definirEstadoPro,
                                definirEstadoUtPro, definirBooleano)

LOG_PROCESO_PROACTIVA = dict()
campanasPorEjecutivos = dict()
listaEstadoRetencion = listaEstadoRetencionProactiva()
campanasValidasRetencion = dict()
filaSalidaTxt = dict()
mantieneSuProducto = listaEstadoRetencion.get('Mantiene su producto')
realizaPagoEnLinea = listaEstadoRetencion.get('Realiza pago en línea')
realizaActivacion = listaEstadoRetencion.get('Realiza Activación PAC/PAT')
listaValidaRetencion = {mantieneSuProducto: mantieneSuProducto, realizaPagoEnLinea: realizaPagoEnLinea, realizaActivacion: realizaActivacion}

def getEstado(celdaEstado):
    listaContactado = {'Pendiente': 1, 'Terminado con Exito': 2 , 'Terminado sin Exito': 3}
    if listaContactado.get(str(celdaEstado.value)):
        return listaContactado[celdaEstado.value]
    elif str(celdaEstado.value) == 'Sin Gestion':
        return 0
    else:
        return False

def getInversaEstado(estado):
    listaContactado = {1: 'Pendiente', 2: 'Terminado con Exito', 3: 'Terminado sin Exito'}
    if listaContactado.get(int(estado)):
        return listaContactado[estado]
    elif int(estado) == 0:
        return 'Sin Gestion'
    else:
        return False

def getEstadoUt(celdaEstadoUt, listaEstadoUt):
    estadoUt = celdaEstadoUt
    if listaEstadoUt.get(str(estadoUt.value)):
        return listaEstadoUt[estadoUt.value]
    elif estadoUt.value is None:
        return 0
    else:
        return False

def validarEstadoRetencion(estado):
    valorEstado = 0
    if listaEstadoRetencion.get(estado):
        valorEstado = listaEstadoRetencion.get(estado)
    return valorEstado

def aprobarCobranza(nroPolizaCertificado, fechaCierre, nroPolizaCliente, fecUltimoPago):
    if fecUltimoPago is not None and fechaCierre is not None:
        if nroPolizaCliente == nroPolizaCertificado and  fecUltimoPago >= fechaCierre:
            return 1
    elif nroPolizaCliente == nroPolizaCertificado and fecUltimoPago is None:
        return 1
    return 0

def aprobarActivacion(estadoMandato, fechaMandato, fechaCierre):
    estadoMandatoValido = PROACTIVA_CONFIG_XLSX['ESTADO_MANDATO_VALIDO']
    estadoActivacion = 0
    if estadoMandatoValido.get(estadoMandato):
        estadoActivacion = 1
    return estadoActivacion

def insertarPeriodoCampanaEjecutivos(campanasEjecutivos: dict, fechaProceso):
    try:
        db = conectorDB()
        cursor = db.cursor()
        ejecutivosExistentes = periodoCampanasProactiva(fechaProceso)
        periodoEjecutivos = convertirListaProactiva(campanasEjecutivos, ejecutivosExistentes, fechaProceso)
        if len(periodoEjecutivos) > 0:
            sql = """INSERT INTO proactiva_campanas_periodo_ejecutivos (id_ejecutivo, periodo) VALUES (?, ?);"""
            cursor.executemany(sql, periodoEjecutivos)
            db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al insertar Periodo de Ejecutivos: %s' % (e))
    finally:
        cursor.close()
        db.close()

def insertarCampanaEjecutivos(campanasEjecutivos: dict, fechaProceso):
    try:
        db = conectorDB()
        camapanasPeriodoEjecutivos = periodoCampanasProactiva(fechaProceso)
        campanasPorPeriodo = []
        cursor = db.cursor()

        for valores in campanasEjecutivos.values():
            for polizas in valores.values():
                idEjecutivo = polizas['ID_EMPLEADO']
                if camapanasPeriodoEjecutivos.get(idEjecutivo):
                    campanasPorPeriodo += setearCampanasProactiva(polizas, camapanasPeriodoEjecutivos[idEjecutivo]['ID'])

        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'INSERTAR_CAMPANAS_EJECUTIVOS': '-----------------------------------------------------' })
        campanasExistentes = CampanasPorPeriodoProactiva(fechaProceso)
        if limpiarTablaCamapanasEjecutivos(fechaProceso) and limpiarPolizasReliquidasAnterior(fechaProceso):
            LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'LIMPIAR_CAMAPAÑAS_EJECUTIVOS': 'EliminarCampanaEjecutivos;Se eliminaron {0} Camapaña(s) existentes'.format(campanasExistentes)})

        if len(campanasPorPeriodo) > 0:
            sql = """INSERT INTO proactiva_campanas_ejecutivos (id_periodo_ejecutivo, numero_poliza, campana_id, nombre_campana, estado_retencion, cobranza_pro, pacpat_pro, cobranza_rel_pro, pacpat_rel_pro, estado_pro, estado_ut_pro, fecha_cierre, reliquidacion, numero_poliza_certificado, polizas_campana, nombre_campana_completo, fecha_creacion, fecha_expiracion_coret, fecha_ultimo_pago, fecha_mandato, estado_mandato, fecha_reliquidacion) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL);"""
            cursor.executemany(sql, campanasPorPeriodo)
            db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al insertar Campañas de Ejecutivos del Periodo: %s' % (e))
    finally:
        cursor.close()
        db.close()

def limpiarTablaCamapanasEjecutivos(fechaProceso):
    try:
        db = conectorDB()
        cursor = db.cursor()

        sql = """DELETE FROM proactiva_campanas_ejecutivos WHERE id_periodo_ejecutivo IN (SELECT id FROM proactiva_campanas_periodo_ejecutivos WHERE periodo = ?);"""
        cursor.execute(sql, fechaProceso)
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al eliminar Campañas de Ejecutivos existentes: %s' % (e))
    finally:
        cursor.close()
        db.close()

def limpiarPolizasReliquidasAnterior(fechaProceso):
    try:
        db = conectorDB()
        cursor = db.cursor()
        mesAnterior = fechaMesAnterior(fechaProceso)

        sql = """UPDATE proactiva_campanas_ejecutivos SET fecha_reliquidacion = NULL FROM proactiva_campanas_ejecutivos WHERE  id_periodo_ejecutivo IN (SELECT id FROM proactiva_campanas_periodo_ejecutivos WHERE periodo = ?) AND reliquidacion = 1 AND fecha_reliquidacion = ?;"""
        cursor.execute(sql, (mesAnterior, fechaProceso))
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al actualizar Campañas por Reliquidar del mes anterior: %s' % (e))
    finally:
        cursor.close()
        db.close()

def agregarCampanasPorEjecutivo(idEmpleado, pk, valoresCampanas: dict):

    if campanasPorEjecutivos.get(idEmpleado):
        campanasPorEjecutivos[idEmpleado].setdefault(pk, valoresCampanas)
    else:
        campanasPorEjecutivos[idEmpleado] = {pk: valoresCampanas}
    return 1

def polizasReliquidadas(periodo, complementoCliente):
    mesAnterior = formatearFechaMesAnterior(periodo)
    fechaIncioMes = primerDiaMes(periodo)
    polizasParaReliquidar = buscarPolizasReliquidar(mesAnterior)
    polizasAprobadaReliquidar = dict()

    for poliza in polizasParaReliquidar.values():

        numeroPoliza = poliza['POLIZA']

        if not complementoCliente.get(numeroPoliza):
            mensaje = 'PolizaReliquidacion;No existe la POLIZA en el COMPLEMENTO_CLIENTE;{0}'.format(numeroPoliza)
            LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_COMPLEMENTO_CLIENTE': mensaje})
            continue

        numeroPolizaCertificado = poliza['NUMERO_POLIZA_CERTIFICADO']
        nombreCampana = poliza['NOMBRE_CAMPANA']
        fechaCierre = poliza['FECHA_CIERRE']
        idEmpleado = poliza['CODIGO_EMPLEADO']
        campanaId = poliza['CAMPAÑA_ID']
        numeroPolizaCliente = complementoCliente[numeroPoliza]['NRO_CERT']
        fecUltimoPago = complementoCliente[numeroPoliza]['FEC_ULT_PAG']
        pk = '{0}_{1}_{2}'.format(campanaId, idEmpleado, numeroPoliza)
        cobranzaRelPro = 0
        pacpatRelPro = 0

        if poliza['COBRANZA_RL_PRO'] > 0:
            cobranzaRelPro = aprobarCobranza(numeroPolizaCertificado, fechaCierre, numeroPolizaCliente, fecUltimoPago)
            if cobranzaRelPro == 0:
                mensaje = 'PolizaReliquidacion;No cumple condicion de retencion COBRANZA para Reliquidacion;%s' % (numeroPoliza)
                LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'PROCESO_COBRANZA_RELIQUIDACION': mensaje})

        if poliza['PACPAT_RL_PRO'] > 0:
            estadoMandato = complementoCliente[numeroPoliza]['ESTADO_MANDATO']
            fecMandato = complementoCliente[numeroPoliza]['FECHA_MANDATO']

            if estadoMandato is not None:
                pacpatRelPro = aprobarActivacion(str(estadoMandato).upper(), fecMandato, fechaCierre)
                mensajeValidacion = 'MANDATOS'
            else:
                pacpatRelPro = aprobarCobranza(numeroPolizaCertificado, fechaCierre, numeroPolizaCliente, fecUltimoPago)
                mensajeValidacion = 'MANDATOS/COBRANZA'
    
            if pacpatRelPro == 0:
                mensaje = 'PolizaReliquidacion;No cumple condicion de retencion {0} para Reliquidacion;{1}'.format(mensajeValidacion, numeroPoliza)
                LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'PROCESO_ACTIVACION_RELIQUIDACION': mensaje})

        if cobranzaRelPro > 0 or pacpatRelPro > 0:
            polizasAprobadaReliquidar[pk] = {'COBRANZA_REL_PRO': cobranzaRelPro, 'PACPAT_REL_PRO': pacpatRelPro, 'CODIGO_EMPLEADO': idEmpleado, 'CAMPAÑA_ID': campanaId, 'NOMBRE_CAMAPANA': nombreCampana, 'POLIZA': numeroPoliza}
    
    if len(polizasAprobadaReliquidar) > 0:
        actualizarPolizasReliquidadas(polizasAprobadaReliquidar, fechaIncioMes, mesAnterior)
        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZAS_RELIQUIDADS': 'PolizasReliquidadas;Existen {0} polizas para reliquidar del mes {1}'.format(len(polizasAprobadaReliquidar), mesAnterior)})
    else:
        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZAS_RELIQUIDADS': 'PolizasReliquidadas;No Existen polizas para reliquidar del periodo: {0}'.format(mesAnterior)})

    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZAS_RELIQUIDADS': '-----------------------------------------------------' })
    return polizasAprobadaReliquidar

def actualizarPolizasReliquidadas(polizasReliquidadas, fechaProceso, mesAnterior):
    try:
        db = conectorDB()
        cursor = db.cursor()
        polizasParaActualizar = []

        for valores in polizasReliquidadas.values():
            polizasParaActualizar.append([fechaProceso, valores['POLIZA'], valores['CAMPAÑA_ID'], valores['CODIGO_EMPLEADO'], mesAnterior])

        sql = """UPDATE proactiva_campanas_ejecutivos SET fecha_reliquidacion = ? WHERE numero_poliza = ? AND campana_id = ? AND id_periodo_ejecutivo = (SELECT id FROM proactiva_campanas_periodo_ejecutivos WHERE id_ejecutivo = ? AND periodo = ?) AND reliquidacion = 1 AND fecha_reliquidacion IS NULL;"""
        cursor.executemany(sql,polizasParaActualizar)
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise Exception('Error al actualizar polizas reliquidadas: %s' % (e))
    finally:
        cursor.close()
        db.close()

def definirEstadoRetencionCampana(nombreCampana):
    if nombreCampana == 'CO RET - Cobranza':
        estadoRetencionValido = mantieneSuProducto
    elif nombreCampana == 'CO RET - Fallo en Instalacion de Mandato':
        estadoRetencionValido = realizaActivacion
    return estadoRetencionValido

def actualizarEstadosRetencionAntiguos(campanasValidasRetencion, campanaId, idEmpleado):
    for llave, valores in campanasValidasRetencion[campanaId].items():
        estadoRetencionOriginal = campanasValidasRetencion[campanaId][llave]['ESTADO_ORIGINAL']
        # if estadoRetencionOriginal is None or estadoRetencionOriginal == '':
        #     estadoRetencionValido = campanasValidasRetencion[campanaId][llave]['ESTADO_CAMBIO']
        if estadoRetencionOriginal is not None:
            if not listaValidaRetencion.get(estadoRetencionOriginal):
                filaSalidaTxt[llave]['ESTADO_RETENCION_PRO'] = estadoRetencionOriginal
                filaSalidaTxt[llave]['COBRANZA_PRO'] = 0
                filaSalidaTxt[llave]['PACPAT_PRO'] = 0
                if campanasPorEjecutivos[idEmpleado].get(llave):
                    campanasPorEjecutivos[idEmpleado].pop(llave)
    return True

def estadoRetencionOriginalValido(campanasValidasRetencion, campanaId):
    cambioAprobado = False
    if campanasValidasRetencion.get(campanaId):
        for llave, valores in campanasValidasRetencion[campanaId].items():
            estadoRetencionOriginal = campanasValidasRetencion[campanaId][llave]['ESTADO_ORIGINAL']
            if listaValidaRetencion.get(estadoRetencionOriginal):
                cambioAprobado = True
    return cambioAprobado

def validarEstadoRetencionCampanaDuplicada(campanasValidasRetencion, campanaId, nombreCampana, estadoRetencion, estado, pk):
    # estadoRetencionValidoExiste = estadoRetencionOriginalValido(campanasValidasRetencion, campanaId)
    estadoRetencionValido = definirEstadoRetencionCampana(nombreCampana)
    if not campanasValidasRetencion.get(campanaId):
        campanasValidasRetencion[campanaId] = {pk: {'PK':pk, 'CAMPAÑA_ID': campanaId, 'ESTADO_ORIGINAL': listaEstadoRetencion.get(estadoRetencion), 'ESTADO_CAMBIO': estadoRetencionValido}}
    elif campanasValidasRetencion.get(campanaId):
        valores = {'PK':pk, 'CAMPAÑA_ID': campanaId, 'ESTADO_ORIGINAL': listaEstadoRetencion.get(estadoRetencion), 'ESTADO_CAMBIO': estadoRetencionValido}
        campanasValidasRetencion[campanaId].setdefault(pk, valores)
    # else:
    #     print(estadoRetencion)
    #     if estadoRetencion is not None or estadoRetencion != '':
    #         estadoRetencionValido = listaEstadoRetencion.get(estadoRetencion)
    #     else:
    #         estadoRetencionValido = validarEstadoRetencion(estado)
    return estadoRetencionValido

def validarRetencionesPolizas(valoresEntrada: dict, complementoCliente: dict):

    estadoRetencion = valoresEntrada['ESTADO_RETENCION']
    nombreCampana = valoresEntrada['NOMBRE_CAMPAÑA']
    numeroPoliza = valoresEntrada['NUMERO_POLIZA']
    idEmpleado = valoresEntrada['ID_EMPLEADO']
    numeroPolizaCertificado = valoresEntrada['NUMERO_POLIZA_CERTIFICADO']
    fechaCierre = valoresEntrada['FECHA_CIERRE']
    campanaId = valoresEntrada['CAMPAÑA_ID']
    estadoValido = valoresEntrada['ESTADO_VALIDO']
    estadoUtValido = valoresEntrada['ESTADO_VALIDOUT']
    celdaNroPoliza = valoresEntrada['CELDA_NROPOLIZA']
    fechaCreacion = valoresEntrada['FECHA_CREACION']
    polizasCampanas = valoresEntrada['POLIZAS_CAMPANA']
    fechaExpiracionCoret = valoresEntrada['FECHA_EXPIRACION_CORET']
    cobranzaPro = 0
    pacpatPro = 0
    seReliquida = 0
    polizaNoAprobada = 0
    controPolizaNoAprobada = False

    if nombreCampana == 'CO RET - Cobranza' and estadoRetencion == realizaActivacion:
        cobranzaPro = 1
        pacpatPro = 1
    elif nombreCampana == 'CO RET - Fallo en Instalacion de Mandato' and estadoRetencion == mantieneSuProducto:
        cobranzaPro = 1
        pacpatPro = 1
    elif nombreCampana == 'CO RET - Cobranza' and estadoRetencion == mantieneSuProducto or nombreCampana == 'CO RET - Cobranza' and estadoRetencion == realizaPagoEnLinea:
        cobranzaPro = 1
    elif nombreCampana == 'CO RET - Fallo en Instalacion de Mandato' and estadoRetencion == realizaPagoEnLinea:
        cobranzaPro = 1
    elif nombreCampana == 'CO RET - Fallo en Instalacion de Mandato' and estadoRetencion == realizaActivacion:
        pacpatPro = 1

    fecUltimoPago = complementoCliente[numeroPoliza]['FEC_ULT_PAG']
    estadoMandato = complementoCliente[numeroPoliza]['ESTADO_MANDATO']
    fecMandato = complementoCliente[numeroPoliza]['FECHA_MANDATO']
    numeroPolizaCliente = complementoCliente[numeroPoliza]['NRO_CERT']
    pk2 = '{0}_{1}_{2}'.format(campanaId, idEmpleado, numeroPoliza)

    valoresPoliza = {'ID_EMPLEADO': idEmpleado, 'NUMERO_POLIZA': numeroPoliza, 'CAMPAÑA_ID': campanaId, 'NOMBRE_CAMPAÑA': nombreCampana, 'ESTADO_RETENCION': estadoRetencion, 'RETENCION_COBRANZA': 0, 'RETENCION_ACTIVACION': 0, 'RETENCION_RL_COBRANZA': 0, 'RETENCION_RL_ACTIVACION': 0, 'ESTADO_VALIDO': estadoValido, 'ESTADO_VALIDOUT': estadoUtValido, 'FECHA_CIERRE': fechaCierre, 'RELIQUIDACION': seReliquida, 'NUMERO_POLIZA_CERTIFICADO': numeroPolizaCertificado, 'FECHA_CREACION': fechaCreacion, 'POLIZAS_CAMPANA': polizasCampanas, 'FECHA_EXPIRACION_CORET': fechaExpiracionCoret, 'FECHA_ULTIMO_PAGO': fecUltimoPago, 'ESTADO_MANDATO': estadoMandato, 'FECHA_MANDATO': fecMandato}
    agregarCampanasPorEjecutivo(idEmpleado, pk2, valoresPoliza)

    if cobranzaPro > 0:
        cobranzaPro = aprobarCobranza(numeroPolizaCertificado, fechaCierre, numeroPolizaCliente, fecUltimoPago)
        if cobranzaPro == 0:
            if campanasPorEjecutivos[idEmpleado].get(pk2):
                campanasPorEjecutivos[idEmpleado][pk2]['RETENCION_RL_COBRANZA'] = 1
                campanasPorEjecutivos[idEmpleado][pk2]['RELIQUIDACION'] = 1
                controPolizaNoAprobada = True

            celdaCoordenada = setearCelda2(celdaNroPoliza, 0)
            mensaje = '%s;Poliza no cumple condicion de retencion COBRANZA;%s' % (celdaCoordenada, numeroPoliza)
            LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'PROCESO_COBRANZA': mensaje})
        else:
            campanasPorEjecutivos[idEmpleado][pk2]['RETENCION_COBRANZA'] = cobranzaPro

    if pacpatPro > 0:
    
        if estadoMandato is not None:
            pacpatPro = aprobarActivacion(str(estadoMandato).upper(), fecMandato, fechaCierre)
            mensajeValidacion = 'MANDATOS'
        else:
            pacpatPro = aprobarCobranza(numeroPolizaCertificado, fechaCierre, numeroPolizaCliente, fecUltimoPago)
            mensajeValidacion = 'MANDATOS/COBRANZA'
    
        if pacpatPro == 0:
            if campanasPorEjecutivos[idEmpleado].get(pk2):
                campanasPorEjecutivos[idEmpleado][pk2]['RETENCION_RL_ACTIVACION'] = 1
                campanasPorEjecutivos[idEmpleado][pk2]['RELIQUIDACION'] = 1
                controPolizaNoAprobada = True

            celdaCoordenada = setearCelda2(celdaNroPoliza, 0)
            mensaje = '{0};Poliza no cumple condicion de retencion {1};{2}'.format(celdaCoordenada, mensajeValidacion, numeroPoliza)
            LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'PROCESO_COBRANZA': mensaje})
        else:
            campanasPorEjecutivos[idEmpleado][pk2]['RETENCION_ACTIVACION'] = pacpatPro

    if controPolizaNoAprobada:
        polizaNoAprobada = 1
    return cobranzaPro, pacpatPro, polizaNoAprobada
        
def leerArchivoProactiva(archivoEntrada, periodo, archivoComplementoCliente):
    try:
        encabezadoXls = PROACTIVA_CONFIG_XLSX['ENCABEZADO_XLSX']
        encabezadoTxt = PROACTIVA_CONFIG_XLSX['ENCABEZADO_TXT']
        encabezadoReliquidacionesTxt = PROACTIVA_CONFIG_XLSX['ENCABEZADO_RELIQUIDACIONES']
        columna = PROACTIVA_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        coordenadaEcabezado = PROACTIVA_CONFIG_XLSX['COORDENADA_ENCABEZADO']
        xls = load_workbook(archivoEntrada, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]

        archivo_correcto = validarEncabezadoXlsx(hoja[coordenadaEcabezado], encabezadoXls, archivoEntrada)
        if type(archivo_correcto) is not dict:

            fechaIncioMes = primerDiaMes(periodo)
            fechaFinMes = ultimoDiaMes(periodo)
            fechaFinMesSiguiente = mesSiguienteUltimoDia(periodo)
            listaConsiderarRetencion = {'Mantiene su producto': mantieneSuProducto, 'Realiza pago en línea': realizaPagoEnLinea, 'Realiza Activación PAC/PAT': realizaActivacion}
            i = 0
            polizasNoAprobadas = 0
            cantidadCampanasValidas = 0
            dataXlsx = dict()
            complementoCliente = extraerComplementoCliente(len(LOG_PROCESO_PROACTIVA), archivoComplementoCliente)
            LOG_PROCESO_PROACTIVA.update(LOG_COMPLEMENTO_CLIENTE)
            ejecutivosExistentesDb = buscarRutEjecutivosDb(fechaFinMes, fechaIncioMes)
            listaEstadoContactado = listaEstadoUtContacto()
            listaEstadoRetencionTexto = estadoRetencionProDesc()
            listaEstadoUtTexto = listaEstadoUtDesc()
            listaEstadoUt = listaEstadoUtAll()
            LOG_PROCESO_PROACTIVA.setdefault('INICIO_LECTURA_PROACTIVA', {len(LOG_PROCESO_PROACTIVA)+1: '-----------------------------------------------------' })
            LOG_PROCESO_PROACTIVA.setdefault('ENCABEZADO_PROACTIVA', {len(LOG_PROCESO_PROACTIVA)+1: 'Encabezado del Archivo: %s OK' % archivoEntrada})
            LOG_PROCESO_PROACTIVA.setdefault('INICIO_CELDAS_PROACTIVA', {len(LOG_PROCESO_PROACTIVA)+1: 'Iniciando lectura de Celdas del Archivo: %s' % archivoEntrada})

            for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo Proactiva' , unit=' Fila'):

                i += 1
                if i >= 2:

                    nombreCampana = str(fila[columna['NOMBRE_DE_CAMPAÑA']].value)
                    codigoEjecutivo = str(fila[columna['ID_EMPLEADO']].value)
                    estado = str(fila[columna['ESTADO']].value)
                    estadoRetencion = fila[columna['ESTADO_RETENCION']].value
                    campanaId = str(fila[columna['CAMAPAÑA_ID']].value)
                    estadoUltimaTarea = fila[columna['ESTADO_ULTIMA_TAREA']].value
                    numeroPoliza, numeroPolizaCertificado = formatearNumeroPoliza(fila[columna['NRO_POLIZA']].value)
                    polizasCampanas = str(fila[columna['POLIZAS_CAMPANAS']].value)
                    pk = '{0}_{1}_{2}'.format(campanaId, codigoEjecutivo, numeroPoliza)

                    if numeroPoliza is None:
                        celdaCoordenada = setearCelda2(fila[0:columna['NRO_POLIZA']+1], len(fila[0:columna['NRO_POLIZA']])-1, i)
                        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_NULO': '%s;Numero de poliza NULL;%s' % (celdaCoordenada, numeroPoliza)})
                        continue

                    fechaCreacion = setearFechaCelda(fila[columna['FECHA_CREACION']])
                    fechaCierre = None
                    fechaExpiracionCoret = None
                    if fila[columna['EXPIRACION_CORET']].value is not None and str(fila[columna['EXPIRACION_CORET']].value) != '':
                        fechaExpiracionCoret = setearFechaCelda(fila[columna['EXPIRACION_CORET']])
                    if fila[columna['FECHA_CIERRE']].value is not None and str(fila[columna['FECHA_CIERRE']].value) != '':
                        fechaCierre = setearFechaCelda(fila[columna['FECHA_CIERRE']])

                    estadoValido = getEstado(fila[columna['ESTADO']])
                    estadoUtValido = getEstadoUt(fila[columna['ESTADO_ULTIMA_TAREA']], listaEstadoUt)

                    if type(fechaCreacion) is not datetime.date:
                        valorErroneo = str(fila[columna['FECHA_CREACION']].value)
                        celdaCoordenada = setearCelda2(fila[0:columna['FECHA_CREACION']+1], len(fila[0:columna['FECHA_CREACION']])-1, i)
                        mensaje = '%s;FECHA_CREACION no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'FECHA_CREACION': mensaje})
                        continue

                    if estado != 'Sin Gestion' and  type(fechaCierre) is not datetime.date:
                        valorErroneo = str(fila[columna['FECHA_CIERRE']].value)
                        celdaCoordenada = setearCelda2(fila[0:columna['FECHA_CIERRE']+1], len(fila[0:columna['FECHA_CIERRE']])-1, i)
                        mensaje = '%s;FECHA_CIERRE no es una fecha valida;%s' % (celdaCoordenada,valorErroneo)
                        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'FECHA_CIERRE': mensaje})
                        continue

                    if estado == 'Sin Gestion' and type(fechaExpiracionCoret) is not datetime.date:
                        valorErroneo = str(fila[columna['EXPIRACION_CORET']].value)
                        celdaCoordenada = setearCelda2(fila[0:columna['EXPIRACION_CORET']+1], len(fila[0:columna['EXPIRACION_CORET']])-1, i)
                        mensaje = '%s;FECHA_EXPIRACION_CORET no es una fecha valida;%s' % (celdaCoordenada, valorErroneo)
                        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'FECHA_EXPIRACION': mensaje})
                        continue

                    if type(estadoValido) is not int:
                        valorErroneo = str(fila[columna['ESTADO']].value)
                        celdaCoordenada = setearCelda2(fila[0:columna['ESTADO']+1], len(fila[0:columna['ESTADO']])-1, i)
                        mensaje = '%s;No existe Estado;%s' % (celdaCoordenada, valorErroneo)
                        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'ERROR_ESTADO': mensaje})
                        continue

                    if type(estadoUtValido) is not int:
                        valorErroneo = str(fila[columna['ESTADO_ULTIMA_TAREA']].value)
                        celdaCoordenada = setearCelda2(fila[0:columna['ESTADO_ULTIMA_TAREA']+1], len(fila[0:columna['ESTADO_ULTIMA_TAREA']])-1, i)
                        mensaje = '%s;No existe EstadoUltimaTarea;%s' % (celdaCoordenada, valorErroneo)
                        LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'ERROR_ESTADOUT': mensaje})
                        continue

                    if estado != 'Sin Gestion' and fechaCierre >= fechaIncioMes and fechaCierre <= fechaFinMes or estado == 'Sin Gestion' and fechaExpiracionCoret >= fechaIncioMes and fechaExpiracionCoret <= fechaFinMesSiguiente:

                        if ejecutivosExistentesDb.get(codigoEjecutivo):
                            idEmpleado = ejecutivosExistentesDb[codigoEjecutivo]['ID_EMPLEADO']
                        else:
                            celdaCoordenada = setearCelda2(fila[0:columna['ID_EMPLEADO']+1], len(fila[0:columna['ID_EMPLEADO']])-1, i)
                            mensaje = '%s;Ejecutivo no existe en la DB;%s' % (celdaCoordenada, codigoEjecutivo)
                            LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'RUT_NO_EXISTE': mensaje})
                            continue
                                

                        repeticionPorCampana = 1
                        if filaSalidaTxt.get(pk):

                            celdaCoordenada = setearCelda2(fila[0:columna['NRO_POLIZA']+1], len(fila[0:columna['NRO_POLIZA']])-1, i)
                            if estado == 'Terminado con Exito':
    
                                if filaSalidaTxt[pk]['ESTADO_PRO'] != 2:
                                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_DUPLICADA': '{0};CAMBIO_POLIZA;ESTADO_ANTERIOR: {1}:NUEVO_VALOR:{2}'.format(celdaCoordenada, getInversaEstado(filaSalidaTxt[pk]['ESTADO_PRO']), estado)})
                                    repeticionPorCampana = filaSalidaTxt[pk]['REPETICIONES'] + 1 
                                    filaSalidaTxt.pop(pk)
                                    dataXlsx.pop(pk)
                                elif listaConsiderarRetencion.get(estadoRetencion):
                                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_DUPLICADA': '{0};CAMBIO_POLIZA(RetencionDoble);ESTADO_ANTERIOR:({1},{2}):NUEVO_VALOR:({3},{4})'.format(celdaCoordenada, getInversaEstado(filaSalidaTxt[pk]['ESTADO_PRO']), listaEstadoRetencionTexto.get(filaSalidaTxt[pk]['ESTADO_RETENCION_PRO']), estado, estadoRetencion)})
                                    repeticionPorCampana = filaSalidaTxt[pk]['REPETICIONES'] + 1
                                    filaSalidaTxt.pop(pk)
                                    dataXlsx.pop(pk)
                                    if campanasPorEjecutivos[idEmpleado].get(pk):
                                        if campanasPorEjecutivos[idEmpleado][pk]['RETENCION_ACTIVACION'] == 1 or campanasPorEjecutivos[idEmpleado][pk]['RETENCION_RL_COBRANZA'] == 1:
                                            polizasNoAprobadas -= 1
                                        campanasPorEjecutivos[idEmpleado].pop(pk)
                                    cantidadCampanasValidas -= 1
                                else:
                                    mensaje = '{0};POLIZA_DUPLICADA;ELIMINADO({1},{2})_vs_PERMANECE({3},{4})'.format(celdaCoordenada, estado, estadoRetencion, getInversaEstado(filaSalidaTxt[pk]['ESTADO_PRO']), listaEstadoRetencionTexto.get(filaSalidaTxt[pk]['ESTADO_RETENCION_PRO']))
                                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_DUPLICADA': mensaje})
                                    filaSalidaTxt[pk]['REPETICIONES'] += 1
                                    continue

                            elif estado == 'Pendiente' or estado == 'Terminado sin Exito':
                                if filaSalidaTxt[pk]['ESTADO_PRO'] != 2 and listaEstadoContactado.get(estadoUltimaTarea):
                                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_DUPLICADA': '{0};CAMBIO_POLIZA;ESTADO_ANTERIOR:({1},{2}):NUEVO_VALOR:({3},{4})'.format(celdaCoordenada, getInversaEstado(filaSalidaTxt[pk]['ESTADO_PRO']), listaEstadoUtTexto.get(filaSalidaTxt[pk]['ESTADO_UT_PRO']), estado, estadoUltimaTarea)})
                                    repeticionPorCampana = filaSalidaTxt[pk]['REPETICIONES'] + 1 
                                    filaSalidaTxt.pop(pk)
                                    dataXlsx.pop(pk)
                                elif filaSalidaTxt[pk]['ESTADO_PRO'] == 0:
                                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_DUPLICADA': '{0};CAMBIO_POLIZA;ESTADO_ANTERIOR:({1},{2}):NUEVO_VALOR:({3},{4})'.format(celdaCoordenada, getInversaEstado(filaSalidaTxt[pk]['ESTADO_PRO']), listaEstadoUtTexto.get(filaSalidaTxt[pk]['ESTADO_UT_PRO']), estado, estadoUltimaTarea)})
                                    repeticionPorCampana = filaSalidaTxt[pk]['REPETICIONES'] + 1 
                                    filaSalidaTxt.pop(pk)
                                    dataXlsx.pop(pk)
                                else:
                                    mensaje = '{0};POLIZA_DUPLICADA;ELIMINADO({1},{2})_vs_PERMANECE({3},{4})'.format(celdaCoordenada, estado, estadoUltimaTarea, getInversaEstado(filaSalidaTxt[pk]['ESTADO_PRO']), listaEstadoUtTexto.get(filaSalidaTxt[pk]['ESTADO_UT_PRO']))
                                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZA_DUPLICADA': mensaje})
                                    filaSalidaTxt[pk]['REPETICIONES'] += 1
                                    continue

                        cobranzaPro = 0
                        pacpatPro = 0

                        listaConsiderarRetencion = {'Mantiene su producto': mantieneSuProducto, 'Realiza pago en línea': realizaPagoEnLinea, 'Realiza Activación PAC/PAT': realizaActivacion,}
                        estadoRetencionValido = listaConsiderarRetencion.get(estadoRetencion)
                        if estado == 'Terminado con Exito':

                            if estadoRetencionValido is None:
                                estadoRetencionValido = validarEstadoRetencionCampanaDuplicada(campanasValidasRetencion, campanaId, nombreCampana, str(estadoRetencion).strip(), estado, pk)

                            else:
                                if campanasValidasRetencion.get(campanaId):
                                    valores = {'PK':pk, 'CAMPAÑA_ID': campanaId, 'ESTADO_ORIGINAL': estadoRetencionValido, 'ESTADO_CAMBIO': 0}
                                    campanasValidasRetencion[campanaId].setdefault(pk, valores)
                                else:
                                    campanasValidasRetencion[campanaId] = {pk: {'PK':pk, 'CAMPAÑA_ID': campanaId, 'ESTADO_ORIGINAL': estadoRetencionValido, 'ESTADO_CAMBIO': 0}}

                            if complementoCliente.get(numeroPoliza) and listaValidaRetencion.get(estadoRetencionValido):
                                
                                valoresEntrada = {'ESTADO_RETENCION': estadoRetencionValido, 'NOMBRE_CAMPAÑA': nombreCampana, 'NUMERO_POLIZA': numeroPoliza, 'FECHA_CIERRE': fechaCierre, 'ID_EMPLEADO': idEmpleado, 'NUMERO_POLIZA_CERTIFICADO': numeroPolizaCertificado, 'CAMPAÑA_ID': campanaId, 'ESTADO_VALIDO': estadoValido, 'ESTADO_VALIDOUT': estadoUtValido, 'CELDA_NROPOLIZA': fila[columna['NRO_POLIZA']], 'FECHA_CREACION': fechaCreacion, 'POLIZAS_CAMPANA': polizasCampanas, 'FECHA_EXPIRACION_CORET': fechaExpiracionCoret}
                                cobranzaPro, pacpatPro, noAprobada = validarRetencionesPolizas(valoresEntrada, complementoCliente)
                                polizasNoAprobadas += noAprobada
                                cantidadCampanasValidas += 1

                        else:
                            if estadoRetencionValido is not None:
                                celdaCoordenada = setearCelda2(fila[0:columna['ESTADO']+1], len(fila[0:columna['ESTADO']])-1, i)
                                mensaje = '{0};ESTADO no corresponde con la RETENCION;{1}/{2}'.format(celdaCoordenada, estado, estadoRetencion)
                                LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'ESTADO_RETENCION_NO_VALIDO': mensaje})

                            if estadoRetencion is not None:
                                estadoRetencionValido = listaEstadoRetencion.get(estadoRetencion)
                            else:
                                estadoRetencionValido = validarEstadoRetencion(estado)

                        filaSalidaTxt[pk] = {'COBRANZA_PRO': cobranzaPro, 'PACPAT_PRO': pacpatPro, 'ESTADO_PRO': estadoValido, 'ESTADO_UT_PRO': estadoUtValido, 'REPETICIONES': repeticionPorCampana, 'ESTADO_RETENCION_PRO': estadoRetencionValido, 'ID_EMPLEADO': idEmpleado, 'CAMPAÑA_ID': campanaId, 'CAMPANA': nombreCampana[0:30].rstrip(), 'POLIZA': numeroPoliza}
                        
                        fecUltimoPago = None
                        estadoMandato = None
                        fecMandato = None
                        if complementoCliente.get(numeroPoliza):
                            fecUltimoPago = complementoCliente[numeroPoliza]['FEC_ULT_PAG']
                            estadoMandato = complementoCliente[numeroPoliza]['ESTADO_MANDATO']
                            fecMandato = complementoCliente[numeroPoliza]['FECHA_MANDATO']

                        dataXlsx[pk] = {'FECHA_CREACION': fechaCreacion, 'NOMBRE_CAMPANA': nombreCampana, 'CODIGO_EMPLEADO': idEmpleado, 'ESTADO_PRO': definirEstadoPro(estadoValido), 'POLIZAS_CAMPANA': polizasCampanas, 'FECHA_CIERRE': fechaCierre, 'POLIZA': str(numeroPoliza), 'FECHA_EXPIRACION': fechaExpiracionCoret, 'ESTADO_RETENCION': definirEstadoRetencionPro(estadoRetencionValido), 'CAMPAÑA_ID': campanaId, 'ESTADO_UT_PRO': definirEstadoUtPro(estadoUtValido), 'FECHA_ULTPAGO': fecUltimoPago, 'ESTADO_MANDATO': estadoMandato, 'FECHA_MANDATO': fecMandato, 'PAGA_COBRANZA': definirBooleano(cobranzaPro), 'PAGA_MANDATO': definirBooleano(pacpatPro)}

                        validarEstadoRetencionExistente = estadoRetencionOriginalValido(campanasValidasRetencion, campanaId)
                        if validarEstadoRetencionExistente:
                            actualizarEstadosRetencionAntiguos(campanasValidasRetencion, campanaId, idEmpleado)

            if insertarPeriodoCampanaEjecutivos(campanasPorEjecutivos, fechaIncioMes):
                if insertarCampanaEjecutivos(campanasPorEjecutivos, fechaIncioMes):
                    mensaje = 'InsertarCampanaEjecutivos;Se insertaron correctamente: %s Campaña(s)' % (cantidadCampanasValidas)
                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'INSERTAR_CAMPANAS_EJECUTIVOS': mensaje})
                    LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'INSERTAR_CAMPANAS_EJECUTIVOS': '-----------------------------------------------------' })

            if polizasNoAprobadas > 0:
                mensaje = 'InsertPolizasReliquidar;Se insertaron correctamente: %s Poliza(s)' % (polizasNoAprobadas)
                LOG_PROCESO_PROACTIVA.setdefault(len(LOG_PROCESO_PROACTIVA)+1, {'POLIZAS_RELIQUIDAR': mensaje})

            polizasReliquidadaTxt = polizasReliquidadas(periodo, complementoCliente)

            LOG_PROCESO_PROACTIVA.setdefault('FIN_CELDAS_PROACTIVA', {len(LOG_PROCESO_PROACTIVA)+1: 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivoEntrada, len(tuple(hoja.rows)))})
            LOG_PROCESO_PROACTIVA.setdefault('PROCESO_PROACTIVA', {len(LOG_PROCESO_PROACTIVA)+1: 'Proceso del Archivo: %s Finalizado' % archivoEntrada})

            return filaSalidaTxt, encabezadoTxt, polizasReliquidadaTxt, encabezadoReliquidacionesTxt, dataXlsx
        else:
            LOG_PROCESO_PROACTIVA.setdefault('ENCABEZADO_PROACTIVA', archivo_correcto)
            raise
    except Exception as e:
        errorMsg = 'Error: %s | %s' % (archivoEntrada, e)
        LOG_PROCESO_PROACTIVA.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_PROACTIVA)+1: errorMsg})
        LOG_PROCESO_PROACTIVA.setdefault('PROCESO_PROACTIVA', {len(LOG_PROCESO_PROACTIVA)+1: 'Error al procesar Archivo: %s' % archivoEntrada})
        return False, False, False, False, False