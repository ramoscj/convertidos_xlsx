from openpyxl import load_workbook
from tqdm import tqdm

from config_xlsx import CAMPANHAS_PRIORITARIAS
from validaciones_texto import setearCelda2

LOG_PROCESO_PRIORITARIAS = dict()

def leerArchivoPrioritarias(archivo):
    try:
        LOG_PROCESO_PRIORITARIAS.setdefault('INICIO_LECTURA_PRIORITARIAS', {len(LOG_PROCESO_PRIORITARIAS)+1: 'Iniciando proceso de lectura del Archivo: %s' % archivo})
        encabezadoXls = CAMPANHAS_PRIORITARIAS['ENCABEZADO_XLSX']
        encabezadoTxt = CAMPANHAS_PRIORITARIAS['ENCABEZADO_TXT']
        columna = CAMPANHAS_PRIORITARIAS['COLUMNAS_PROCESO_XLSX']
        coordenadaEcabezado = CAMPANHAS_PRIORITARIAS['COORDENADA_ENCABEZADO']
        xls = load_workbook(archivo, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        correlativo = 1
        
        LOG_PROCESO_PRIORITARIAS.setdefault(len(LOG_PROCESO_PRIORITARIAS)+1, {'ENCABEZADO_PRIORITARIAS': 'Encabezado del Archivo: %s OK' % archivo})
        filaSalidaXls = dict()
        empleadosLista = dict()
        totalFilas = len(tuple(hoja.iter_rows(min_row=1, min_col=1)))
        LOG_PROCESO_PRIORITARIAS.setdefault(len(LOG_PROCESO_PRIORITARIAS)+1, {'INICIO_CELDAS_PRIORITARIAS': 'Iniciando lectura de Celdas del Archivo: %s' % archivo})

        i = 1
        for fila in tqdm(iterable=hoja.iter_rows(min_row=1, min_col=1), total = totalFilas, desc='Leyendo PrioritariasCRO' , unit=' Fila'):
        # for fila in hoja.iter_rows(min_row=3, min_col=1):
            campanhasPrioritarias = fila[columna['CAMPANA']].value
            prioritaria = fila[columna['PRIORITARIA']].value

            if campanhasPrioritarias is None or campanhasPrioritarias == '':
                valorVacio = '%s;El valor de Campaña es NULL;%s' % (setearCelda2(fila[columna['CAMPANA']], 0), campanhasPrioritarias)
                LOG_PROCESO_PRIORITARIAS.setdefault(len(LOG_PROCESO_PRIORITARIAS)+1, {'CAMPAÑA_VACIO': valorVacio})
                continue
            if prioritaria is None or type(prioritaria) is not int:
                errorRut = '%s;El valor de Prioritario no es valido;%s' % (setearCelda2(fila[columna['PRIORITARIA']], 0), prioritaria)
                LOG_PROCESO_PRIORITARIAS.setdefault(len(LOG_PROCESO_PRIORITARIAS)+1, {'PRIORITARIO_VACIO': errorRut})
                continue
            
            filaSalidaXls[i] = {'CRR': i, 'PRIORITARIA': int(prioritaria), 'CAMPANA': campanhasPrioritarias[0:30]}
            i += 1

        LOG_PROCESO_PRIORITARIAS.setdefault('FIN_CELDAS_PRIORITARIAS', {len(LOG_PROCESO_PRIORITARIAS)+1: 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivo, len(tuple(hoja.rows)))})

        LOG_PROCESO_PRIORITARIAS.setdefault('PROCESO_PRIORITARIAS', {len(LOG_PROCESO_PRIORITARIAS)+1: 'Proceso del Archivo: %s Finalizado' % archivo})
        return filaSalidaXls, encabezadoTxt

    except Exception as e:
        errorMsg = 'Error: %s | %s' % (archivo, e)
        LOG_PROCESO_PRIORITARIAS.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_PRIORITARIAS)+1: errorMsg})
        LOG_PROCESO_PRIORITARIAS.setdefault('PROCESO_PRIORITARIAS', {len(LOG_PROCESO_PRIORITARIAS)+1: 'Error al procesar Archivo: %s' % archivo})
        return False, False

# x = leerArchivoPrioritarias('CRO/INPUTS/Campanas_Prioritarias_CRO.xlsx')
# print(LOG_PROCESO_PRIORITARIAS)
