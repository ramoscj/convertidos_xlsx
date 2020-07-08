from openpyxl import load_workbook
from tqdm import tqdm
from tqdm.auto import tqdm
import csv

import os.path
import sys
import datetime

def leer_xlsx(archivo):
    try:
        data_txt = []
        encabezado = ['LPATTR_PER_RES', 'LLAVEA', 'LPATTR_COD_POLI', 'LPATTR_COD_ORIGEN', 'TIPO', 'LPATTR_COD_STAT', 'NRO_POLIZA', 'ESTADOPOLIZA', 'FECHAINICIOVIGENCIA', 'RUT_CRO', 'NOMBRE_CRO', 'FECHAPROCESO', 'CONSIDERAR_FUGA']
        # read_only=True para leer el archivo y consumir menos recursos
        # data_only=True para leer el valor de las celdas que tiene formulas
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]
        i = 0
        archivo_correcto = False
        for columna in hoja['A1:M1']:
            for celda in columna:
                dato = str(celda.value)
                if dato.upper() == encabezado[i]:
                    archivo_correcto = True
                else:
                    print('Error columna: %s' % i)
                    archivo_correcto = False
                i += 1
        if archivo_correcto:
            cantidad_columnas = len(tuple(hoja.rows))
            i = 1
            for columna in tqdm(iterable=hoja.rows, total = cantidad_columnas, desc='Leyendo DATA' , unit='Row'):
                data_xls = []
                control_data = False
                for celda in range(0, len(columna)):
                    data_xls.append(i)
                    considerar_fuga = str(columna[12].value)
                    if celda == 4:
                        tipo = str(columna[4].value)
                        # PREGUNTAR CUANDO CONSIDERAR_FUGA ESTE VACIO
                        if tipo.upper() == 'FUGA' and considerar_fuga.upper() != 'NO':
                            data_xls.append(1)
                            control_data = True
                        else:
                            control_data = False
                            data_xls = []
                    if control_data and celda == 5:
                        lpattr_cod_stat = str(columna[5].value)
                        # PREGUNTAR CUANDO CONSIDERAR_FUGA ESTE VACIO
                        if lpattr_cod_stat.upper() == 'NVIG' and considerar_fuga.upper() != 'NO':
                            data_xls.append(1)
                        else:
                            control_data = False
                            data_xls = []
                    if control_data and celda == 9:
                        rut_cro = str(columna[9].value)
                        if rut_cro is not None:
                            data_xls.append(rut_cro)
                        else:
                            data_xls = []
                if len(data_xls) != 0:
                    data_txt.append(data_xls)
                    i += 1
        else:
            print('Error el archivo presenta incosistencias en el encabezado')
        return data_txt
    except Exception as e:
        print('Error al leer archivo: %s | %s' % (archivo, e))

def escribir_txt(archivo_salida, data_txt):
    try:
        with open(archivo_salida, 'w', newline='') as txt:
            writer = csv.writer(txt, delimiter=';')
            for data in tqdm(iterable=data_txt, total = len(data_txt), desc='Escribiendo DATA', unit='Row'):
                writer.writerow(data)
        return True
    except Exception as e:
        print('Error al escribir archivo: %s | %s' % (archivo_salida, e))

def validar_fecha(f1, f2):
    try:
        if type(datetime.date(int(f1), int(f2), 1)) is datetime.date:
            return True
    except Exception as e:
        print("Error de fecha, formato correcto YYYYMM | %s" % e)

if len(sys.argv) == 3:
    proceso = str(sys.argv[1])
    fecha_x = str(sys.argv[2])
    fecha_a = fecha_x[0:4]
    fecha_b = fecha_x[4:6]
    if proceso.upper() == 'FUGA':
        if validar_fecha(fecha_a, fecha_b):
            archivo = '%s_FUGA_AGENCIA.xlsx' % fecha_x
            archivo_existe = os.path.isfile(archivo)
            salida_txt = 'FUGA%s.txt' % fecha_x
            if archivo_existe:
                print("Archivo: %s encontrado." % archivo)
                print("Iniciando Lectura...")
                data_txt = leer_xlsx(archivo)
                if escribir_txt(salida_txt, data_txt):
                    print("Archivo Creado !!")
                else:
                    print("Error!!")
            else:
                print('Error: Archivo %s no encontrado' % archivo)
    else:
        print('Error: Proceso %s no encontrado' % proceso)
else:
    print("Error: El programa "'"%s"'" necesita dos parametros" % sys.argv[0])










# with tqdm(total=len(my_list)) as pbar:
            #     for dato in data_txt:
            #         pbar.update(1)

# for i in tqdm(range(100001)):
#     print("", end='\r')

# loop = tqdm(total = 5000, position = 0, leave= False)
# for k in range(5000):
#     loop.set_description("Cargando...".format(k))
#     loop.update(1)
# loop.close()

# from tqdm import tqdm
# import requests

# chunk_size = 1024

# url = "http://www.tutorialspoint.com/python3/python_tutorial.pdf"

# req = requests.get(url, stream = True)

# total_size = int(req.headers['content-length'])

# with open("pythontutorial.pdf", "wb") as file:
#     for data in tqdm(iterable=req.iter_content(chunk_size=chunk_size), total = total_size/chunk_size, unit='KB'):
#         file.write(data)

# print("Download Completed !!!")