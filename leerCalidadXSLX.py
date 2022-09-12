from openpyxl import load_workbook
from tqdm import tqdm
import traceback

from config_xlsx import CALIDAD_CONFIG_XLSX
from diccionariosDB import buscarRutEjecutivosDb
from validaciones_texto import (formatearRut, setearCelda2,
                                validarEncabezadoXlsx, ultimoDiaMes, primerDiaMes)

LOG_PROCESO_CALIDAD = dict()

def leerArchivoCalidad(archivo, periodo):
    try:
        LOG_PROCESO_CALIDAD.setdefault('INICIO_LECTURA', {len(LOG_PROCESO_CALIDAD)+1: 'Iniciando proceso de lectura del Archivo: %s' % archivo})
        encabezadoTxt = CALIDAD_CONFIG_XLSX['ENCABEZADO_TXT']
        celda = CALIDAD_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]

        LOG_PROCESO_CALIDAD.setdefault(len(LOG_PROCESO_CALIDAD)+1, {'ENCABEZADO_ARCHIVO': 'Encabezado del Archivo: %s OK' % archivo})
        i = 0
        correlativo = 1
        filaSalidaXls = dict()
        ejecutivosExistentesDb = buscarRutEjecutivosDb(ultimoDiaMes(periodo), primerDiaMes(periodo))

        for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo Calidad' , unit=' Fila'):

            if i >= 2 and fila[celda['ID_EMPLEADO']].value is not None:
                idEmpleado = str(fila[celda['ID_EMPLEADO']].value)
                if ejecutivosExistentesDb.get(idEmpleado):
                    calidad = int(float(fila[celda['CALIDAD']].value)*100)
                    filaSalidaXls[idEmpleado] = {'CRR': correlativo, 'CALIDAD': calidad, 'ID_EMPLEADO': idEmpleado}
                    correlativo += 1
                else:
                    errorRut = '%s;No existe Ejecutivo;%s' % (setearCelda2(fila[celda['ID_EMPLEADO']], 0), idEmpleado)
                    LOG_PROCESO_CALIDAD.setdefault(len(LOG_PROCESO_CALIDAD)+1, {'EJECUTIVO_NO_EXISTE_%s' % i: errorRut})
            i += 1
        LOG_PROCESO_CALIDAD.setdefault(len(LOG_PROCESO_CALIDAD)+1, {'FIN_LECTURA_ARCHIVO': 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivo, len(tuple(hoja.rows)))})
        LOG_PROCESO_CALIDAD.setdefault(len(LOG_PROCESO_CALIDAD)+1, {'FIN_PROCESO': 'Proceso del Archivo: %s Finalizado' % archivo})
        return filaSalidaXls, encabezadoTxt

    except Exception as e:
        LOG_PROCESO_CALIDAD.setdefault(len(LOG_PROCESO_CALIDAD)+1, {'LECTURA_ARCHIVO_ERROR': traceback.format_exc()})
        LOG_PROCESO_CALIDAD.setdefault(len(LOG_PROCESO_CALIDAD)+1, {'PROCESO_ARCHIVO': 'Error al procesar Archivo: %s' % archivo})
        return False, False

# leerArchivoCalidad('INPUTS/202005_Calidad_CRO.xlsx', '202005')
# print(LOG_PROCESO_CALIDAD)
