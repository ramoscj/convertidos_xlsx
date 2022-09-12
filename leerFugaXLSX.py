import datetime
import traceback

from openpyxl import load_workbook
from tqdm import tqdm

from config_xlsx import FUGA_CONFIG_XLSX, PATH_XLSX
from diccionariosDB import buscarRutEjecutivosDb
from validaciones_texto import (formatearRut, setearCelda, validaFechaCelda,
                                validarEncabezadoXlsx, ultimoDiaMes, primerDiaMes)

LOG_PROCESO_FUGA = dict()

def validarFugaStock(correlativo, tipo, lpattrCodStat, idEmpleado, unidad):
    rutNuevoFuga = dict()
    rutNuevoFuga["CRR"] = correlativo
    if tipo == 'FUGA':
        rutNuevoFuga["FUGA"] = 1
    else:
        rutNuevoFuga["FUGA"] = 0
    if lpattrCodStat != 'NVIG':
        rutNuevoFuga["STOCK"] = 1
    else:
        rutNuevoFuga["STOCK"] = 0
    rutNuevoFuga["ID_EMPLEADO"] = idEmpleado
    rutNuevoFuga["UNIDAD"] = unidad
    return rutNuevoFuga

def existeRut(tipo, lpattrCodStat, rutExistenteFuga):
    if tipo == 'FUGA':
        rutExistenteFuga["FUGA"] += 1
    elif lpattrCodStat != 'NVIG':
        rutExistenteFuga["STOCK"] += 1
    return rutExistenteFuga

def leerArchivoFuga(archivo, periodo):
    try:
        LOG_PROCESO_FUGA.setdefault('INICIO_LECTURA_FUGA', {len(LOG_PROCESO_FUGA)+1: 'Iniciando proceso de lectura del Archivo: %s' % archivo})
        encabezadoFugaTxt = FUGA_CONFIG_XLSX['ENCABEZADO_FUGA_TXT']
        columna = FUGA_CONFIG_XLSX['COLUMNAS_PROCESO_XLSX']
        xls = load_workbook(archivo, read_only=True, data_only=True)
        nombre_hoja = xls.sheetnames
        hoja = xls[nombre_hoja[0]]

        LOG_PROCESO_FUGA.setdefault(len(LOG_PROCESO_FUGA)+1, {'ENCABEZADO_FUGA': 'Encabezado del Archivo: %s OK' % archivo})
        i = 0
        correlativo = 1
        filaSalidaFugaXls = dict()
        filaSalidaStockXls = dict()
        ejecutivosExistentesDb = buscarRutEjecutivosDb(ultimoDiaMes(periodo), primerDiaMes(periodo))
        LOG_PROCESO_FUGA.setdefault(len(LOG_PROCESO_FUGA)+1, {'INICIO_CELDAS_FUGA': 'Iniciando lectura de Celdas del Archivo: %s' % archivo})
        
        for fila in tqdm(iterable=hoja.rows, total = len(tuple(hoja.rows)), desc='Leyendo FugaCRO' , unit=' Fila'):

            if i >= 1:
                fechaLpattrs = validaFechaCelda(fila[columna['LPATTR_PER_RES']])
                if type(fechaLpattrs) is str:
                        LOG_PROCESO_FUGA.setdefault(len(LOG_PROCESO_FUGA)+1, {'FECHA_CREACION': fechaLpattrs})
                        continue
                if periodo == str(fechaLpattrs.value) and fila[columna['ID_EMPLEADO']].value is not None:

                    idEmpleado = str(fila[columna['ID_EMPLEADO']].value)
                    tipo = str(fila[columna['TIPO']].value).upper()
                    #considerarFuga = str(fila[columna['CONSIDERAR_FUGA']].value).upper()
                    lpattrCodStat = str(fila[columna['LPATTR_COD_STAT']].value).upper()

                    if ejecutivosExistentesDb.get(idEmpleado):
                        unidad = ejecutivosExistentesDb[idEmpleado]['PLATAFORMA']
                        if filaSalidaFugaXls.get(idEmpleado):
                            filaSalidaFugaXls[idEmpleado] = existeRut(tipo, lpattrCodStat, filaSalidaFugaXls[idEmpleado])
                        else:
                            filaSalidaFugaXls[idEmpleado] = validarFugaStock(correlativo, tipo, lpattrCodStat, idEmpleado, unidad)
                            correlativo += 1
                    else:
                        errorRut = 'Celda%s - No existe Ejecutivo: %s' % (setearCelda(fila[columna['ID_EMPLEADO']]), idEmpleado)
                        LOG_PROCESO_FUGA.setdefault('EJECUTIVO_NO_EXISTE_%s' % i, {len(LOG_PROCESO_FUGA)+1: errorRut})
            i += 1

        LOG_PROCESO_FUGA.setdefault('FIN_CELDAS_FUGA', {len(LOG_PROCESO_FUGA)+1: 'Lectura de Celdas del Archivo: %s Finalizada - %s filas' % (archivo, len(tuple(hoja.rows)))})
        LOG_PROCESO_FUGA.setdefault('PROCESO_FUGA', {len(LOG_PROCESO_FUGA)+1: 'Proceso del Archivo: %s Finalizado' % archivo})
        return filaSalidaFugaXls, encabezadoFugaTxt

    except Exception as e:
        # errorMsg = 'Error: %s | %s' % (archivo, e)
        LOG_PROCESO_FUGA.setdefault('LECTURA_ARCHIVO', {len(LOG_PROCESO_FUGA)+1: traceback.format_exc()})
        LOG_PROCESO_FUGA.setdefault('PROCESO_FUGA', {len(LOG_PROCESO_FUGA)+1: 'Error al procesar Archivo: %s' % archivo})
        return False, False
        # raise
